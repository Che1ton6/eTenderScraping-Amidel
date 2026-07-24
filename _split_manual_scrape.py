#!/usr/bin/env python3
"""Split Thando's 19 July manual scrape into two per-batch folders based on
each tender's PUBLICATION_DATE:
  - (T) 13-15 July 2026
  - (M) 16-19 July 2026

Produces in each folder:
  - <batch> Manual Scrapped.xlsx   (detail rows for that window)
  - <batch> number of tenders scrapped manual.xlsx  (per-source counts)
"""
import io, sys, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from BatchProcessor import SOURCES, _xl_match

MANUAL_ROOT = os.path.join(HERE, 'data', 'Manual Scrapes')
SRC_DETAIL  = os.path.join(MANUAL_ROOT, '19 July 2026', '19-07-2026 Manual Scrapped.xlsx')
SRC_SUMMARY = os.path.join(MANUAL_ROOT, '19 July 2026', '19-07-2026 number of tenders scrapped manual.xlsx')

# Batch definitions
BATCHES = [
    {'label': '(T) 13-15 July 2026', 'from': date(2026, 7, 13), 'to': date(2026, 7, 15)},
    {'label': '(M) 16-19 July 2026', 'from': date(2026, 7, 16), 'to': date(2026, 7, 19)},
]

# ── Load Thando's detail file ────────────────────────────────────────────────
detail = pd.read_excel(SRC_DETAIL)
# Drop blank/header rows (row 0 was all-NaN in the original)
detail = detail.dropna(how='all').reset_index(drop=True)
print(f"Loaded {len(detail)} detail rows from Thando's file\n")

def _parse_pub(v):
    if pd.isna(v):
        return None
    if hasattr(v, 'date'):
        return v.date() if hasattr(v, 'year') else None
    s = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            return pd.to_datetime(s, format=fmt).date()
        except Exception:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).date()
    except Exception:
        return None


detail['__pub_date'] = detail['PUBLICATION_DATE'].apply(_parse_pub)

def _match_source_label(dept):
    """Return the SOURCES label that matches the given DEPARTMENT string, or None."""
    if not isinstance(dept, str):
        return None
    for label, pattern in SOURCES:
        if pattern is None:
            continue
        if _xl_match(dept, pattern):
            return label
    # Manual fallbacks — Thando abbreviates some
    dl = dept.lower().strip()
    aliases = {
        'sita': 'SITA', 'labour': 'DEL', 'ecdpw': 'EC DPW', 'jpc': 'JPC',
        'siu': 'SIU', 'gdoh': 'GDoH', 'gpl': 'GPL', 'raf': 'RAF',
        'winnie madikizela mandela': 'Winnie Mandela LM',
        'winnie mandela lm': 'Winnie Mandela LM',
        'nmb': 'Nelson Mandela Bay MM', 'buffalo city': 'Buffalo City MM',
    }
    return aliases.get(dl)


detail['__source_label'] = detail['DEPARTMENT '].apply(_match_source_label) \
    if 'DEPARTMENT ' in detail.columns else detail['DEPARTMENT'].apply(_match_source_label)


HDR_FONT = Font(bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1C3880')
TOTAL_FILL = PatternFill('solid', fgColor='F5A000')
_thin = Side(style='thin', color='C0C0C0')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _write_detail(out_path, sub_df):
    """Write a per-batch detail xlsx mirroring Thando's schema."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    # Use the same columns as Thando's file, minus internal marker columns
    cols = [c for c in sub_df.columns if not str(c).startswith('__') and not str(c).startswith('Unnamed')]
    for c_idx, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    for r_idx, row in enumerate(sub_df[cols].itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            v = '' if pd.isna(val) else val
            ws.cell(row=r_idx, column=c_idx, value=v).border = BORDER
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = min(max_len + 2, 50)
    ws.freeze_panes = 'A2'
    wb.save(out_path)


def _write_summary(out_path, counts_by_source, batch_label, total):
    """Write per-source count summary mirroring Thando's second file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    # Header row
    ws.cell(row=1, column=1, value='Source').font = HDR_FONT
    ws.cell(row=1, column=1).fill = HDR_FILL
    ws.cell(row=1, column=1).border = BORDER
    ws.cell(row=1, column=2, value='No').font = HDR_FONT
    ws.cell(row=1, column=2).fill = HDR_FILL
    ws.cell(row=1, column=2).border = BORDER

    # Report Date row
    ws.cell(row=2, column=1, value='Report Date').border = BORDER
    ws.cell(row=2, column=2, value=batch_label).border = BORDER

    # Per-source rows
    r = 3
    for label, _ in SOURCES:
        if label == 'eTenders (All)':
            continue
        ws.cell(row=r, column=1, value=label).border = BORDER
        ws.cell(row=r, column=2, value=counts_by_source.get(label, 0)).border = BORDER
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        r += 1
    # Total row
    ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value=total).font = Font(bold=True)
    ws.cell(row=r, column=2).fill = TOTAL_FILL
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 8
    wb.save(out_path)


# ── Process each batch ───────────────────────────────────────────────────────
for batch in BATCHES:
    label, df_from, df_to = batch['label'], batch['from'], batch['to']
    print(f"=== {label} ({df_from} to {df_to}) ===")

    sub = detail[detail['__pub_date'].apply(
        lambda d: d is not None and df_from <= d <= df_to)].copy()
    print(f"  {len(sub)} manual tenders fall in this window")
    for _, r in sub.iterrows():
        print(f"    PUB={r['__pub_date']}  SRC={r.get('TENDER_SOURCE'):<28}  ID={str(r.get('TENDER_ID'))[:26]}")

    # Per-source counts based on the SOURCES labels
    counts = {}
    for _, r in sub.iterrows():
        lbl = r.get('__source_label')
        if lbl:
            counts[lbl] = counts.get(lbl, 0) + 1

    # Prepare output folder
    out_dir = os.path.join(MANUAL_ROOT, label)
    os.makedirs(out_dir, exist_ok=True)
    detail_path = os.path.join(out_dir, f"{label} Manual Scrapped.xlsx")
    summary_path = os.path.join(out_dir, f"{label} number of tenders scrapped manual.xlsx")

    _write_detail(detail_path, sub)
    _write_summary(summary_path, counts, label, len(sub))
    print(f"  wrote: {os.path.basename(detail_path)}")
    print(f"  wrote: {os.path.basename(summary_path)}")
    print()

print("Done.")
