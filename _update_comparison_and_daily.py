#!/usr/bin/env python3
"""
Update the Automated_vs_Manual_Comparison spreadsheet and DAILYTENDER_REPORT_3.0.

Steps:
  1. Load existing Automated_vs_Manual_Comparison.xlsx from Sales/ folder.
  2. Preserve historical batch columns as-is.
  3. Add two new batch column blocks for (T) 13-15 July and (M) 16-19 July.
     - Automated = per-source count from our end product
     - Manual    = per-source count from Thando's split files
     - Missed    = tenders (by normalised TENDER_ID) in Manual but not in Auto
  4. Save the updated file to  Sales/Thando and Tenders/Manual Scraping/
     (moving it out of Sales/ root).
  5. Append the 16 manual tenders to DAILYTENDER_REPORT_3.0 (Scraper).xlsx
     TENDER_DATABASE sheet.
"""
import io, sys, os, shutil, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from BatchProcessor import SOURCES, _xl_match, _normalize_tender_id

USER = os.path.expanduser('~')
SALES = os.path.join(USER, 'OneDrive - Amidel (Pty) Ltd', 'Documents', 'Sales')
SRC_COMPARE = os.path.join(SALES, 'Automated_vs_Manual_Comparison.xlsx')
DEST_DIR = os.path.join(SALES, 'Thando and Tenders', 'Manual Scraping')
DEST_COMPARE = os.path.join(DEST_DIR, 'Automated_vs_Manual_Comparison.xlsx')

DAILY = os.path.join(SALES, 'Thando and Tenders', 'DAILYTENDER_REPORT_3.0 (Scraper).xlsx')

# Our automated batch folders — end products
AUTO_ROOT = os.path.join(HERE, 'data')

BATCHES = [
    {
        'label':      '(T) 13-15 July 2026',
        'col_label':  '13 - 15 July 2026',
        'auto_ep':    os.path.join(AUTO_ROOT, 'etenders.gov.za', '(T) 13-15 July 2026', 'end product', 'RFQ_and_ICT_Checker_(13 - 15 July).xlsx'),
        'manual_dtl': os.path.join(AUTO_ROOT, 'Manual Scrapes', '(T) 13-15 July 2026', '(T) 13-15 July 2026 Manual Scrapped.xlsx'),
    },
    {
        'label':      '(M) 16-19 July 2026',
        'col_label':  '16 - 19 July 2026',
        'auto_ep':    os.path.join(AUTO_ROOT, 'etenders.gov.za', '(M) 16-19 July 2026', 'end product', 'RFQ_and_ICT_Checker_(16 - 19 July).xlsx'),
        'manual_dtl': os.path.join(AUTO_ROOT, 'Manual Scrapes', '(M) 16-19 July 2026', '(M) 16-19 July 2026 Manual Scrapped.xlsx'),
    },
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def _norm(tid: str) -> str:
    if not tid or (isinstance(tid, float) and pd.isna(tid)):
        return ''
    return _normalize_tender_id(str(tid).strip())

def _match_source(dept: str):
    """Return the SOURCES label whose pattern matches this DEPARTMENT."""
    if not isinstance(dept, str):
        return None
    for label, pattern in SOURCES:
        if pattern is None:
            continue
        if _xl_match(dept, pattern):
            return label
    # Manual alias map for Thando's shortened names
    dl = dept.strip().lower()
    aliases = {
        'sita': 'SITA', 'labour': 'DEL', 'ecdpw': 'EC DPW', 'jpc': 'JPC',
        'siu': 'SIU', 'gdoh': 'GDoH', 'gpl': 'GPL', 'raf': 'RAF',
        'winnie madikizela mandela': 'Winnie Mandela LM',
    }
    return aliases.get(dl)

def _analyse(auto_df: pd.DataFrame, manual_df: pd.DataFrame):
    """Return per-source dict: {source: (auto_count, manual_count, missed_count)}.
       Missed = # of manual tenders (by normalised TENDER_ID) not in auto for that source.
    """
    auto_dept_col = 'DEPARTMENT' if 'DEPARTMENT' in auto_df.columns else 'DEPARTMENT '
    manual_dept_col = 'DEPARTMENT' if 'DEPARTMENT' in manual_df.columns else 'DEPARTMENT '

    auto_df = auto_df.copy()
    auto_df['__source'] = auto_df[auto_dept_col].apply(_match_source)
    auto_df['__nid']    = auto_df['TENDER_ID'].apply(_norm)

    manual_df = manual_df.copy()
    manual_df['__source'] = manual_df[manual_dept_col].apply(_match_source)
    manual_df['__nid']    = manual_df['TENDER_ID'].apply(_norm)

    result = {}
    for label, _ in SOURCES:
        if label == 'eTenders (All)':
            continue
        auto_sub = auto_df[auto_df['__source'] == label]
        manual_sub = manual_df[manual_df['__source'] == label]
        auto_ids = set(auto_sub['__nid']) - {''}
        manual_ids = set(manual_sub['__nid']) - {''}
        missed = manual_ids - auto_ids
        # Also count manual rows with blank TENDER_ID that don't have any auto row
        # with matching DEPARTMENT+description
        blank_manual = manual_sub[manual_sub['__nid'] == '']
        for _, r in blank_manual.iterrows():
            desc = str(r.get('TENDER_DESCRIPTION') or '').strip()
            if not desc:
                continue
            if not any(str(a.get('TENDER_DESCRIPTION') or '').strip().startswith(desc[:30])
                       for _, a in auto_sub.iterrows()):
                missed.add(f'__blank_{desc[:20]}')
        result[label] = (len(auto_sub), len(manual_sub), len(missed))
    return result


# ── 1. Read existing comparison spreadsheet ──────────────────────────────────

print(f"Reading existing comparison: {os.path.basename(SRC_COMPARE)}")
existing = pd.read_excel(SRC_COMPARE, sheet_name='Comparison', header=None)
print(f"  {existing.shape[0]} rows × {existing.shape[1]} cols")

# Extract historical data: row 0 = batch labels, row 1 = sub-headers,
# rows 2+ = per-source counts. Preserve verbatim.
historical_headers   = existing.iloc[0].tolist()
historical_subheads  = existing.iloc[1].tolist()
historical_rows      = existing.iloc[2:].values.tolist()  # includes TOTAL

# ── 2. Compute new batches' analysis ──────────────────────────────────────────

new_batches_data = []
for batch in BATCHES:
    print(f"\n=== {batch['label']} ===")
    try:
        auto_df = pd.read_excel(batch['auto_ep'], sheet_name='Tender Data')
        print(f"  Auto end product: {len(auto_df)} rows")
    except Exception as e:
        print(f"  ERROR loading auto EP: {e}")
        auto_df = pd.DataFrame()
    try:
        manual_df = pd.read_excel(batch['manual_dtl'])
        print(f"  Manual detail:    {len(manual_df)} rows")
    except Exception as e:
        print(f"  ERROR loading manual: {e}")
        manual_df = pd.DataFrame()

    counts = _analyse(auto_df, manual_df) if len(auto_df) or len(manual_df) else {}
    total_auto   = sum(v[0] for v in counts.values())
    total_manual = sum(v[1] for v in counts.values())
    total_missed = sum(v[2] for v in counts.values())
    print(f"  TOTAL: auto={total_auto}, manual={total_manual}, missed={total_missed}")
    new_batches_data.append({
        'col_label': batch['col_label'],
        'counts':    counts,
        'total':     (total_auto, total_manual, total_missed),
    })


# ── 3. Write the updated comparison workbook ─────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = 'Comparison'

HDR_FONT  = Font(bold=True, color='FFFFFF')
HDR_FILL  = PatternFill('solid', fgColor='1C3880')
SUB_FILL  = PatternFill('solid', fgColor='D9E1F2')
NEW_FILL  = PatternFill('solid', fgColor='FFF2CC')  # highlight new-batch columns
TOTAL_FILL = PatternFill('solid', fgColor='F5A000')
_thin = Side(style='thin', color='C0C0C0')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Row 1 headers
ws.cell(row=1, column=1, value='Source').fill = HDR_FILL
ws.cell(row=1, column=1).font = HDR_FONT
ws.cell(row=1, column=1).border = BORDER
ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

# Existing batch headers preserved
col = 2
existing_batch_labels = []
for i in range(1, existing.shape[1], 3):
    v = existing.iloc[0, i]
    if isinstance(v, str) and v.strip():
        existing_batch_labels.append((v, col))
        for j in range(3):
            cell = ws.cell(row=1, column=col + j, value=v if j == 0 else '')
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        col += 3

# New batch headers
for nb in new_batches_data:
    for j in range(3):
        cell = ws.cell(row=1, column=col + j, value=nb['col_label'] if j == 0 else '')
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
    col += 3

# Row 2 sub-headers (Automated / Manual / Missed)
ws.cell(row=2, column=1, value='').border = BORDER
total_cols = col - 1  # last populated col
sub_col = 2
while sub_col <= total_cols:
    for j, h in enumerate(('Automated', 'Manual', 'Missed')):
        cell = ws.cell(row=2, column=sub_col + j, value=h)
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    sub_col += 3

# Merge existing batch header cells
for lbl, start_col in existing_batch_labels:
    try:
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=start_col + 2)
    except Exception:
        pass

# Data rows — use full SOURCES list (skip eTenders (All))
row = 3
existing_source_col_a = existing.iloc[2:, 0].astype(str).str.strip().tolist()
existing_data_by_src = {}
for r_idx, src in enumerate(existing_source_col_a):
    existing_data_by_src[src] = existing.iloc[2 + r_idx].tolist()

for label, _ in SOURCES:
    if label == 'eTenders (All)':
        continue
    is_total = False
    ws.cell(row=row, column=1, value=label).border = BORDER
    ws.cell(row=row, column=1).font = Font(bold=True)

    # Fill existing batch cells from historical data (if present)
    existing_row = existing_data_by_src.get(label)
    col_cursor = 2
    if existing_row is not None:
        for i in range(1, len(existing_row)):
            v = existing_row[i]
            if pd.notna(v):
                ws.cell(row=row, column=col_cursor, value=v).border = BORDER
                ws.cell(row=row, column=col_cursor).alignment = Alignment(horizontal='center')
            else:
                ws.cell(row=row, column=col_cursor, value=0).border = BORDER
                ws.cell(row=row, column=col_cursor).alignment = Alignment(horizontal='center')
            col_cursor += 1
    else:
        # Source is new to comparison (e.g. JPC, JOSHCO, SIU, FSCA); fill 0
        for i in range(1, existing.shape[1]):
            ws.cell(row=row, column=col_cursor, value=0).border = BORDER
            ws.cell(row=row, column=col_cursor).alignment = Alignment(horizontal='center')
            col_cursor += 1

    # Now write new batch cells (Auto / Manual / Missed for each new batch)
    for nb in new_batches_data:
        a, m, ms = nb['counts'].get(label, (0, 0, 0))
        for j, v in enumerate((a, m, ms)):
            cell = ws.cell(row=row, column=col_cursor + j, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
            cell.fill = NEW_FILL
        col_cursor += 3
    row += 1

# TOTAL row
ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
ws.cell(row=row, column=1).border = BORDER
col_cursor = 2

# For existing batches, preserve TOTAL row from source (may have blanks)
total_source_row = None
for r_idx, src in enumerate(existing_source_col_a):
    if src == 'TOTAL':
        total_source_row = existing.iloc[2 + r_idx].tolist()
        break

if total_source_row is not None:
    for i in range(1, len(total_source_row)):
        v = total_source_row[i]
        if pd.notna(v):
            ws.cell(row=row, column=col_cursor, value=v)
        # leave blank if source had blank
        col_cursor += 1
else:
    col_cursor = 2 + existing.shape[1] - 1

# New batches totals
for nb in new_batches_data:
    a, m, ms = nb['total']
    for j, v in enumerate((a, m, ms)):
        cell = ws.cell(row=row, column=col_cursor + j, value=v)
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    col_cursor += 3

# Column widths
ws.column_dimensions['A'].width = 24
for c in range(2, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(c)].width = 10
ws.freeze_panes = 'B3'

# Save to destination
os.makedirs(DEST_DIR, exist_ok=True)
wb.save(DEST_COMPARE)
print(f"\nSaved updated comparison: {DEST_COMPARE}")

# Move: delete the original from Sales root (comment out if you want to keep it)
if os.path.exists(SRC_COMPARE):
    os.remove(SRC_COMPARE)
    print(f"Removed original: {SRC_COMPARE}")


# ── 4. Append manual tenders to DAILYTENDER_REPORT_3.0 TENDER_DATABASE ──────

print(f"\n=== Updating DAILYTENDER_REPORT_3.0 (Scraper).xlsx ===")

all_manual = []
for batch in BATCHES:
    try:
        df = pd.read_excel(batch['manual_dtl'])
        all_manual.append(df)
    except Exception as e:
        print(f"  skip {batch['label']}: {e}")
all_manual = pd.concat(all_manual, ignore_index=True) if all_manual else pd.DataFrame()
print(f"  Combined manual tenders to append: {len(all_manual)}")

if len(all_manual) and os.path.exists(DAILY):
    # Backup
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    bak = f"{os.path.splitext(DAILY)[0]}_pre_manual_backup_{ts}.xlsx"
    shutil.copy2(DAILY, bak)
    print(f"  backup: {os.path.basename(bak)}")

    # Open the workbook and add rows to TENDER_DATABASE
    wb_daily = load_workbook(DAILY)
    if 'TENDER_DATABASE' in wb_daily.sheetnames:
        ws_daily = wb_daily['TENDER_DATABASE']
        # Determine header row (row 1)
        headers = [ws_daily.cell(row=1, column=c).value for c in range(1, ws_daily.max_column + 1)]
        # Find last used data row
        last_row = ws_daily.max_row
        # Skip rows that are entirely blank (avoid appending after tail blanks)
        # Just append at max_row + 1
        insert_at = last_row + 1

        # Standard mapping: match manual columns to headers by name.
        for i, mr in all_manual.iterrows():
            for c_idx, h in enumerate(headers, 1):
                if not h:
                    continue
                # Try to find matching column in manual row
                for candidate in (h, h.strip(), h.upper(), h.lower(), h.strip().upper()):
                    if candidate in mr.index:
                        v = mr[candidate]
                        if pd.notna(v):
                            ws_daily.cell(row=insert_at, column=c_idx, value=v)
                        break
                else:
                    # Try DEPARTMENT/DEPARTMENT  variants
                    if h.strip().upper() in ('DEPARTMENT', 'DEPARTMENT '):
                        for cand in ('DEPARTMENT ', 'DEPARTMENT'):
                            if cand in mr.index and pd.notna(mr[cand]):
                                ws_daily.cell(row=insert_at, column=c_idx, value=mr[cand])
                                break
            insert_at += 1
        wb_daily.save(DAILY)
        print(f"  appended {len(all_manual)} rows to TENDER_DATABASE (starting row {last_row + 1})")
    else:
        print(f"  TENDER_DATABASE sheet not found — skipping")

print("\nDone.")
