#!/usr/bin/env python3
"""
Tender Analysis Generator
Produces a styled "Tender Analysis.xlsx" inside the batch folder.
One sheet per tracked source; each row classifies the tender for Amidel
and explains missing fields.
"""

import os
import logging
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from BatchProcessor import SOURCES


# ── Amidel scope definition ───────────────────────────────────────────────────

_YES_CATEGORIES = {
    "information service activities",
    "information and communication",
}

_YES_KEYWORDS = [
    "software", "website", "web application", "web app", "mobile application",
    "mobile app", "application development", "app development", "digital platform",
    "online platform", "information system", "management information system",
    "database", "data analytics", "data management", "ict", "information technology",
    "information communication", "cybersecurity", "cyber security", "network",
    "cloud", "system development", "system integration", "erp", "crm", "portal",
    "e-service", "eservice", "digital service",
    "security service", "security services", "security provider",
    "security guard", "guarding service", "manned guarding",
    "close protection", "armed response", "security personnel",
    "cctv", "surveillance", "access control", "biometric",
    "intrusion detection", "alarm system", "security system",
    "physical security", "security solution",
]

_DEBATABLE_KEYWORDS = [
    "monitoring system", "smart meter", "smart system",
    "electronic system", "tracking system", "tracking solution",
    "communication system", "technology solution", "automation",
    "perimeter security", "security fencing", "electric fence",
    "remote monitoring", "telemetry", "scada",
]

_NO_KEYWORDS = [
    "medical", "dental", "optometry", "optician", "eye care", "health clinic",
    "pharmaceutical", "medicine", "clinical",
    "construction", "civil works", "building works", "roads", "road works",
    "bridge", "infrastructure", "renovation", "refurbishment", "maintenance works",
    "fuel", "diesel", "petrol", "lubricant", "oil supply",
    "grass", "lawn", "garden", "landscaping", "mowing", "horticulture",
    "cleaning", "janitorial", "hygiene services", "pest control",
    "food", "catering", "beverage", "meal", "kitchen",
    "transport", "vehicle", "fleet", "bus", "taxi",
    "printing", "stationery", "office supply",
    "furniture", "office furniture",
    "waste management", "landfill", "refuse", "environmental audit",
    "rope access", "mechanical works", "plumbing", "hvac",
    "air conditioning", "electrical installation",
    "occupational hygiene", "ergonomic",
    "legal service", "accounting service", "audit service",
    "supply and delivery of", "supply, delivery",
    "grass cutting", "tree felling", "tree cutting",
    "stipend", "bursary", "scholarship",
    "insurance", "actuarial",
    "translation", "interpretation service",
]

_NO_CATEGORIES = {
    "supplies: medical",
    "supplies: fuel",
    "specialised construction activities",
    "building of complete constructions or parts thereof",
}

_SECURITY_SERVICES_CATEGORY = "services: functional (including cleaning and security services)"


def _classify(tender: dict) -> tuple[str, str]:
    """Returns (verdict, explanation). verdict ∈ {YES, NO, Debatable, MAYBE}."""
    desc     = str(tender.get("TENDER_DESCRIPTION") or "").lower()
    category = str(tender.get("CATEGORY") or "").lower().strip()
    dept     = str(tender.get("DEPARTMENT") or "")

    # Category fast-path → NO
    if any(cat in category for cat in _NO_CATEGORIES):
        label = tender.get("TENDER_DESCRIPTION", "")[:80]
        return "NO", f"{label} — outside Amidel scope. Do not submit."

    # Category fast-path → YES (ICT)
    if any(cat in category for cat in _YES_CATEGORIES):
        label = tender.get("TENDER_DESCRIPTION", "")[:80]
        return "YES", f"ICT category — {label}. Assess fit and submit if applicable."

    # "Functional" category needs description check (cleaning vs security)
    if _SECURITY_SERVICES_CATEGORY in category:
        if any(kw in desc for kw in ["security", "guarding", "surveillance", "cctv", "access control"]):
            label = tender.get("TENDER_DESCRIPTION", "")[:80]
            return "YES", f"Security services — {label}. Direct Amidel fit. Submit."
        label = tender.get("TENDER_DESCRIPTION", "")[:80]
        return "NO", f"{label} — cleaning/functional services, outside Amidel scope. Do not submit."

    # Description keyword checks
    yes_match = next((kw for kw in _YES_KEYWORDS if kw in desc), None)
    if yes_match:
        label = tender.get("TENDER_DESCRIPTION", "")[:80]
        if any(kw in desc for kw in ["security", "guard", "cctv", "surveillance", "access control", "biometric"]):
            return "YES", f"Security services match — {label}. Submit."
        return "YES", f"ICT/software match ('{yes_match}') — {label}. Submit."

    no_match = next((kw for kw in _NO_KEYWORDS if kw in desc), None)
    if no_match:
        label = tender.get("TENDER_DESCRIPTION", "")[:80]
        return "NO", f"{label} — outside Amidel scope. Do not submit."

    debatable_match = next((kw for kw in _DEBATABLE_KEYWORDS if kw in desc), None)
    if debatable_match:
        label = tender.get("TENDER_DESCRIPTION", "")[:80]
        return "Debatable", (
            f"{label} — has a technology/monitoring angle but core may be outside Amidel scope. "
            f"Escalate to review."
        )

    # No strong signal — default to NO with description
    label = tender.get("TENDER_DESCRIPTION", "")[:80]
    return "NO", f"{label} — no clear ICT or security angle. Do not submit."


# ── Per-source field availability notes ───────────────────────────────────────
# Maps SOURCE_KEY → {FIELD_NAME: "reason when blank"}
# Fields listed here are STRUCTURALLY unavailable on that source's website.

SOURCE_FIELD_NOTES: dict[str, dict[str, str]] = {
    # ── eTenders sources ──────────────────────────────────────────────────────
    "ETENDERS.GOV.ZA": {},  # Full data available

    # ── EC DPW (PDF/OCR scraper) ──────────────────────────────────────────────
    "ECDPW.GOV.ZA": {
        "CATEGORY":      "EC DPW tender PDFs do not include a category field",
        "ESUBMISSION":   "EC DPW does not use eTenders eSubmission",
        "CLOSING_TIME":  "EC DPW tender PDFs do not always include a closing time",
    },

    # ── Eastern Cape municipalities ───────────────────────────────────────────
    "MATATIELE.GOV.ZA": {
        "TENDER_ID":        "Matatiele website listing does not show tender reference numbers",
        "CLOSING_TIME":     "Matatiele website does not list closing times",
        "CATEGORY":         "Matatiele website does not categorise tenders",
    },
    "NTABANKULU.GOV.ZA": {
        "TENDER_ID":        "Ntabankulu website listing does not show tender reference numbers",
        "CLOSING_DATE":     "Ntabankulu website listing does not show closing dates (only in linked post)",
        "CLOSING_TIME":     "Ntabankulu website does not list closing times",
        "CATEGORY":         "Ntabankulu website does not categorise tenders",
    },
    "UMZIMVUBU.GOV.ZA": {
        "TENDER_ID":        "Umzimvubu website does not list tender reference numbers",
        "PUBLICATION_DATE": "Umzimvubu website does not list publication dates",
        "CLOSING_DATE":     "Umzimvubu website does not list closing dates (detail is inside the PDF)",
        "CLOSING_TIME":     "Umzimvubu website does not list closing times",
        "CATEGORY":         "Umzimvubu website does not categorise tenders",
    },
    "WINNIEMMLM.GOV.ZA": {
        "TENDER_ID":        "Winnie Mandela LM website does not show tender reference numbers",
        "CLOSING_TIME":     "Winnie Mandela LM website does not list closing times",
        "CATEGORY":         "Winnie Mandela LM website does not categorise tenders",
    },
    "MNQUMA.GOV.ZA": {
        "CLOSING_TIME":  "Mnquma website does not list closing times",
        "CATEGORY":      "Mnquma website does not categorise tenders",
        # Publication date and closing date ARE available on individual tender pages
        # and are fetched per-tender — if blank here, the individual page had no date field
    },
    "GREATKEILM.GOV.ZA": {
        "CATEGORY":         "Great Kei LM website does not categorise tenders",
    },
    "AMAHLATHI.GOV.ZA": {
        "TENDER_ID":        "Amahlathi website listing does not show tender reference numbers",
        "PUBLICATION_DATE": "Amahlathi website does not list publication dates",
        "CLOSING_DATE":     "Closing dates not shown in listing — only on each tender's detail page",
        "CATEGORY":         "Amahlathi website does not categorise tenders",
    },
    "BUFFALOCITY.GOV.ZA": {
        "TENDER_ID":        "buffalocity.gov.za is currently offline (IIS placeholder on all paths) — no data available",
        "PUBLICATION_DATE": "buffalocity.gov.za is currently offline — no data available",
        "CLOSING_DATE":     "buffalocity.gov.za is currently offline — no data available",
        "CLOSING_TIME":     "buffalocity.gov.za is currently offline — no data available",
        "CATEGORY":         "buffalocity.gov.za is currently offline — no data available",
    },
    "NELSONMANDELABAY.GOV.ZA": {
        "PUBLICATION_DATE": "Nelson Mandela Bay website does not list publication dates on the tender listing",
        "CLOSING_DATE":     "Nelson Mandela Bay website tender table only shows SCM No., Description, and Tender Fee — closing date is not displayed on the website",
        "CLOSING_TIME":     "Nelson Mandela Bay website does not list closing times",
        "CATEGORY":         "Nelson Mandela Bay website does not categorise tenders",
    },

    # ── Gauteng ───────────────────────────────────────────────────────────────
    "JOSHCO.CO.ZA": {
        "TENDER_ID":        "JOSHCO website does not list tender reference numbers",
        "PUBLICATION_DATE": "JOSHCO website does not list publication dates",
        "CATEGORY":         "JOSHCO website does not categorise tenders",
    },
    "GPL.GOV.ZA": {
        "CLOSING_DATE":     "GPL website listing does not show closing dates",
        "CLOSING_TIME":     "GPL website does not list closing times",
        "CATEGORY":         "GPL website does not categorise tenders",
    },
    "JOHANNESBURGWATER.CO.ZA": {
        "PUBLICATION_DATE": "Johannesburg Water website does not list publication dates",
        "CATEGORY":         "Johannesburg Water does not use eTenders categories",
    },
    "E-TENDERS.GAUTENG.GOV.ZA": {
        "PUBLICATION_DATE": "Gauteng e-Tenders portal does not list publication dates",
        "CATEGORY":         "Gauteng e-Tenders does not use eTenders category codes",
        # Compulsory flag IS extracted from the briefing session date cell
        # (cell contains date, time, and 'Compulsory'/'Non-compulsory' text)
    },

    # ── National departments / agencies ──────────────────────────────────────
    "RAF.CO.ZA": {
        "TENDER_ID":        "RAF portal (SharePoint) requires JavaScript — cannot be scraped automatically; manual monitoring required",
        "PUBLICATION_DATE": "RAF portal (SharePoint) requires JavaScript — cannot be scraped automatically",
        "CLOSING_DATE":     "RAF portal (SharePoint) requires JavaScript — cannot be scraped automatically",
        "CLOSING_TIME":     "RAF portal (SharePoint) requires JavaScript — cannot be scraped automatically",
        "CATEGORY":         "RAF portal (SharePoint) requires JavaScript — cannot be scraped automatically",
    },
    "LABOUR.GOV.ZA": {
        "PUBLICATION_DATE": "DEL (Labour) website does not list publication dates",
        "CATEGORY":         "DEL (Labour) website does not categorise tenders",
    },
    "DIRCO.GOV.ZA": {
        "CATEGORY":         "DIRCO website does not categorise tenders",
    },
    "SIU.ORG.ZA": {
        "PUBLICATION_DATE": "SIU website does not list publication dates",
        "CATEGORY":         "SIU website does not categorise tenders",
    },
    "SITA.CO.ZA": {
        "CATEGORY":         "SITA website does not categorise tenders using eTenders category codes",
    },
    "JUSTICE.GOV.ZA": {
        "TENDER_ID":        "DOJ website only lists contract extensions, not open tenders",
        "PUBLICATION_DATE": "DOJ website only lists contract extensions, not open tenders",
        "CLOSING_DATE":     "DOJ website only lists contract extensions, not open tenders",
        "CATEGORY":         "DOJ website only lists contract extensions, not open tenders",
    },

    # ── Selenium-rendered ─────────────────────────────────────────────────────
    "CITYPOWER.CO.ZA": {
        "CATEGORY":         "City Power website does not use eTenders category codes",
    },
}


def _field_value(tender: dict, field: str, source_key: str,
                 is_briefing_session: bool) -> str:
    """Return display value for an Analysis column cell.
    If the field has no data, always return a non-blank explanation.
    """
    raw = tender.get(field)
    val = str(raw).strip() if raw is not None else ""

    if val in ("", "NaT", "nan", "None", "NaN"):
        val = ""

    if val:
        return val

    # Check source-level structural notes first (explains WHY the field is missing)
    notes = SOURCE_FIELD_NOTES.get(source_key, {})
    if field in notes:
        return notes[field]

    # Everything else — field was not found on the page
    return "Not found on page"


# ── Styling constants ─────────────────────────────────────────────────────────

_TITLE_FILL  = PatternFill("solid", fgColor="1F3864")
_TITLE_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HDR_FILL    = PatternFill("solid", fgColor="305496")
_HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL    = PatternFill("solid", fgColor="D9E1F2")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_VERDICT_FILL = PatternFill("solid", fgColor="D6DCE4")
_VERDICT_FONT = Font(bold=True, size=10)

_YES_FILL      = PatternFill("solid", fgColor="C6EFCE")
_YES_FONT      = Font(bold=True, color="276221")
_NO_FILL       = PatternFill("solid", fgColor="FCE4D6")
_NO_FONT       = Font(bold=True, color="9C3700")
_DEBATABLE_FILL = PatternFill("solid", fgColor="FFEB9C")
_DEBATABLE_FONT = Font(bold=True, color="9C6500")
_MAYBE_FILL    = PatternFill("solid", fgColor="FFEB9C")
_MAYBE_FONT    = Font(bold=True, color="9C6500")

_COL_WIDTHS = [4, 22, 24, 46, 14, 22, 15, 20, 15, 11, 30, 45, 15, 55]
_WRAP       = Alignment(wrap_text=True, vertical="top")
_CENTRE     = Alignment(horizontal="center", vertical="top")

_ANALYSIS_HEADERS = [
    "#", "Tender ID", "Department", "Description", "Province", "Category",
    "Publication Date", "Closing Date", "Briefing Date", "Compulsory",
    "Briefing Venue", "URL", "Should we Submit?", "Explanation",
]

# Fields aligned to _ANALYSIS_HEADERS; None = handled specially in loop
_ANALYSIS_FIELDS = [
    None,                   # col 1:  # row number
    "TENDER_ID",            # col 2
    "DEPARTMENT",           # col 3
    "TENDER_DESCRIPTION",   # col 4
    "PROVINCE",             # col 5
    "CATEGORY",             # col 6
    "PUBLICATION_DATE",     # col 7
    "CLOSING_DATE",         # col 8  (CLOSING_TIME appended)
    "BRIEFING_DATE",        # col 9
    "COMPULSORY_BRIEFING",  # col 10
    "BRIEFING_SESSION_VENUE", # col 11
    "LINK",                 # col 12: URL / tender portal link
    None,                   # col 13: Should we Submit? — left blank for Claude
    None,                   # col 14: Explanation — left blank for Claude
]


def _fmt_date(val) -> str:
    """Format a date value as 'DD Mon YYYY' if parseable, else return as-is."""
    s = str(val).strip()
    if s in ("", "NaT", "nan", "None"):
        return ""
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%-d %b %Y")
        except (ValueError, AttributeError):
            pass
    # Windows doesn't support %-d; use %d and strip leading zero
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d %b %Y").lstrip("0")
        except (ValueError, AttributeError):
            pass
    return s


def _verdict_style(verdict: str) -> tuple:
    if verdict == "YES":
        return _YES_FILL, _YES_FONT
    if verdict == "NO":
        return _NO_FILL, _NO_FONT
    if verdict == "Debatable":
        return _DEBATABLE_FILL, _DEBATABLE_FONT
    return _MAYBE_FILL, _MAYBE_FONT


def _write_analysis_sheet(wb: Workbook, sheet_name: str,
                           source_df: pd.DataFrame, report_date_str: str) -> None:
    """Write one styled analysis sheet for a single source."""
    ws = wb.create_sheet(sheet_name[:31])
    num_cols = len(_ANALYSIS_HEADERS)

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Amidel (Pty) Ltd — Tender Analysis: {sheet_name}   |   "
                               f"Report Date: {report_date_str}")
    title_cell.fill = _TITLE_FILL
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 18

    # ── Row 2: column headers ─────────────────────────────────────────────────
    for col_idx, header in enumerate(_ANALYSIS_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _CENTRE if col_idx in (1, 10) else _WRAP
    ws.row_dimensions[2].height = 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    source_df = source_df.reset_index(drop=True)

    for i, (_, row) in enumerate(source_df.iterrows()):
        xl_row      = i + 3
        row_fill    = _WHITE_FILL if i % 2 == 0 else _ALT_FILL
        source_key  = str(row.get("TENDER_SOURCE") or "")
        has_briefing = str(row.get("IS_THERE_A_BRIEFING_SESSION") or "").strip().lower() == "yes"

        for col_idx, field in enumerate(_ANALYSIS_FIELDS, 1):
            cell = ws.cell(row=xl_row, column=col_idx)
            cell.fill = row_fill
            cell.alignment = _CENTRE if col_idx in (1, 10) else _WRAP

            if col_idx == 1:
                # Row number
                cell.value = i + 1

            elif col_idx in (13, 14):
                # "Should we Submit?" and "Explanation" — left blank for Claude to complete
                cell.value = ""

            elif field == "CLOSING_DATE":
                cd = _fmt_date(row.get("CLOSING_DATE"))
                ct = str(row.get("CLOSING_TIME") or "").strip()
                if ct in ("", "NaT", "nan", "None"):
                    ct = ""
                if cd and ct:
                    cell.value = f"{cd} {ct}"
                elif cd:
                    cell.value = cd
                else:
                    note = _field_value(row.to_dict(), "CLOSING_DATE", source_key, has_briefing)
                    cell.value = note
                    cell.font  = Font(size=9, color="808080", italic=True)

            elif field == "PUBLICATION_DATE":
                val = _fmt_date(row.get("PUBLICATION_DATE"))
                if val:
                    cell.value = val
                else:
                    note = _field_value(row.to_dict(), "PUBLICATION_DATE", source_key, has_briefing)
                    cell.value = note
                    cell.font  = Font(size=9, color="808080", italic=True)

            elif field == "BRIEFING_DATE":
                val = _fmt_date(row.get("BRIEFING_DATE"))
                if val:
                    cell.value = val
                else:
                    cell.value = _field_value(row.to_dict(), field, source_key, has_briefing)
                    cell.font = Font(size=9, color="808080", italic=True)

            elif field == "LINK":
                raw = str(row.get("LINK") or "").strip()
                if raw and raw.startswith("http"):
                    cell.value = raw
                    cell.hyperlink = raw
                    cell.style = "Hyperlink"
                    cell.alignment = _WRAP
                else:
                    cell.value = "Not found on page"
                    cell.font = Font(size=9, color="808080", italic=True)

            elif field is not None:
                raw = str(row.get(field) or "").strip()
                if raw and raw not in ("NaT", "nan", "None", "NaN"):
                    cell.value = raw
                else:
                    # Field is empty — always show an explanation, always grey italic
                    cell.value = _field_value(row.to_dict(), field, source_key, has_briefing)
                    cell.font = Font(size=9, color="808080", italic=True)

        ws.row_dimensions[xl_row].height = 40

    # ── Verdict summary row (blank — for Claude to fill in) ───────────────────
    last_row = len(source_df) + 3
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=num_cols)
    vc = ws.cell(row=last_row, column=1, value="Verdict summary — YES:   |  Debatable:   |  MAYBE:   |  NO:")
    vc.fill  = _VERDICT_FILL
    vc.font  = _VERDICT_FONT
    vc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[last_row].height = 16

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header rows
    ws.freeze_panes = "A3"


# ── XL-match helper (mirrors TenderSummary.py) ───────────────────────────────

def _xl_match(value: str, pattern: str) -> bool:
    if not isinstance(value, str):
        return False
    regex = re.escape(pattern).replace(r"\*", ".*")
    return bool(re.fullmatch(regex, value, re.IGNORECASE))


# ── Public entry point ────────────────────────────────────────────────────────

def create_tender_analysis(df: pd.DataFrame, batch_folder: str,
                           report_date_str: str) -> int:
    """
    Generate Tender Analysis.xlsx in batch_folder.
    One sheet per tracked source with >= 1 tender.
    Returns number of sheets written.
    """
    if df is None or len(df) == 0:
        logging.info("No tenders — skipping Tender Analysis")
        return 0

    dept_series = df["DEPARTMENT"].fillna("").astype(str)
    is_rfq = df["TENDER_TYPE"].fillna("").str.lower() == "request for quotation"

    wb = Workbook()
    wb.remove(wb.active)
    sheets_created = 0

    for label, pattern in SOURCES:
        if pattern is None:
            continue  # skip "eTenders (All)"

        dept_mask = dept_series.apply(lambda v, p=pattern: _xl_match(v, p))
        source_df = df[dept_mask].copy()
        if len(source_df) == 0:
            continue

        # RFQs first, then by closing date
        source_df["_is_rfq"] = is_rfq[source_df.index]
        source_df = source_df.sort_values(
            ["_is_rfq", "CLOSING_DATE"], ascending=[False, True]
        ).drop(columns=["_is_rfq"])

        _write_analysis_sheet(wb, label, source_df, report_date_str)
        logging.info(f"Tender Analysis sheet: {label} ({len(source_df)} tender(s))")
        sheets_created += 1

    if sheets_created == 0:
        logging.info("No tenders matched any tracked source — Tender Analysis not created")
        return 0

    filepath = os.path.join(batch_folder, "Tender Analysis.xlsx")
    wb.save(filepath)
    logging.info(f"Tender Analysis saved: {filepath} ({sheets_created} sheet(s))")
    return sheets_created
