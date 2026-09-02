#!/usr/bin/env python3
"""
Parts 2 & 3: Creates the local batch folder structure, saves per-day Excel
files, builds the end product file (Tender Data + Final tabs), and appends
the new batch column to the equation file.
"""

import re
import os
import logging
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

EQUATION_FILE = os.environ.get("EQUATION_FILE_PATH") or (
    r"C:\Users\CheltonGraham\OneDrive - Amidel (Pty) Ltd"
    r"\Documents\Sales\Sales Auto Hub\Scraping and Reports"
    r"\ICT & RFQ\Old Method\Bhekis conditional Product\RFQ_and_ICT_Equation.xlsx"
)

POWER_BI_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "eTender_PowerBI_Data.xlsx")

# Exact department matching patterns taken from the Final tab Excel formulas.
# (label, dept_pattern) — None means "all rows" (eTenders All).
SOURCES = [
    ("eTenders (All)",        None),
    ("Eskom",                 "ESKOM"),
    ("SITA",                  "State Information Technology Agency"),
    ("Transnet",              "Transnet*"),
    ("JPC",                   "*Joburg Property*"),
    ("Matatiele LM",          "Matatiele Local Municipality"),
    ("Ntabankulu LM",         "Ntabankulu Local Municipality"),
    ("Umzimvubu LM",          "Umzimvubu Local Municipality"),
    ("Winnie Mandela LM",     "Winnie Madikizela*"),
    ("Mnquma LM",             "Mnquma Local Municipality"),
    ("Great Kei LM",          "Great Kei Local Municipality"),
    ("Amahlathi LM",          "Amahlathi Local Municipality"),
    ("Raymond Mhlaba LM",     "Raymond Mhlaba Local Municipality"),
    ("JOSHCO",                "*Johannesburg Social Housing*"),
    ("GPL",                   "Gauteng Provincial Legislature"),
    ("RAF",                   "Road Accident Fund"),
    ("MM Trading Company",    "*Metropolitan Trading*"),
    ("Nelson Mandela Bay MM", "*Nelson Mandela Bay*"),
    ("Buffalo City MM",       "*Buffalo City*"),
    ("EC DPW",                "*Eastern Cape*Public Works*"),
    ("DEL",                   "*Labour*"),
    ("SIU",                   "Special Investigating Unit*"),
    ("GDoH",                  "*Gauteng*Health*"),
    ("DIRCO",                 "*International Relation*"),
    ("DOJ",                   "*Justice*Constitutional*"),
    ("CP JHB",                "City Power*Johannesburg*"),
    ("W JHB",                 "*Johannesburg Water*"),
    ("FSCA",                  "*Financial Sector Conduct Authority*"),
]

ICT_CATEGORIES = ["Information service activities", "Information and communication"]
RFQ_TYPE       = "Request for Quotation"

# Whether we have a dedicated scraper hitting the organ of state's own website.
# Used by All-but-eTenders mode to show the Website column in the Display
# Equation. "-" for eTenders (All) since it's an umbrella row.
WATCHLIST_HAS_SCRAPER = {
    "eTenders (All)":         "-",
    "Eskom":                  "No",   # blocked (Cloudflare 403)
    "SITA":                   "Yes",
    "Transnet":               "No",   # ECONNREFUSED
    "JPC":                    "Yes",
    "Matatiele LM":           "Yes",
    "Ntabankulu LM":          "Yes",
    "Umzimvubu LM":           "Yes",
    "Winnie Mandela LM":      "Yes",
    "Mnquma LM":              "Yes",
    "Great Kei LM":           "Yes",
    "Amahlathi LM":           "Yes",
    "Raymond Mhlaba LM":      "Yes",
    "JOSHCO":                 "Yes",
    "GPL":                    "Yes",
    "RAF":                    "No",   # stub — SharePoint JS-required
    "MM Trading Company":     "No",
    "Nelson Mandela Bay MM":  "Yes",
    "Buffalo City MM":        "Yes",
    "EC DPW":                 "Yes",
    "DEL":                    "Yes",
    "SIU":                    "Yes",
    "GDoH":                   "Yes",
    "DIRCO":                  "Yes",
    "DOJ":                    "No",   # stub — page has no open tenders
    "CP JHB":                 "Yes",
    "W JHB":                  "Yes",
    "FSCA":                   "Yes",
}

TENDER_COLUMNS = [
    "REPORT_DATE", "RECORD_ID", "TENDER_ID", "PUBLICATION_DATE",
    "CLOSING_DATE", "CLOSING_TIME", "TENDER_TYPE", "TENDER_DESCRIPTION",
    "TENDER_SOURCE", "DEPARTMENT", "PROVINCE",
    "ESUBMISSION", "CATEGORY", "IS_THERE_A_BRIEFING_SESSION",
    "BRIEFING_DATE", "COMPULSORY_BRIEFING", "BRIEFING_SESSION_VENUE", "LINK", "SOE",
    "COST_OF_SALES_ESTIMATE", "CAPABILITY_AVAILABLE", "CAPABILITY_GROUP", "REQUIREMENTS",
    "DUPLICATED", "INGESTION_METHOD",
]
# End product and daily batch files exclude briefing fields (Tender Summary only)
BATCH_COLUMNS = [c for c in TENDER_COLUMNS if c != "BRIEFING_DATE"]
DATE_COLUMNS = {"REPORT_DATE", "PUBLICATION_DATE", "CLOSING_DATE", "BRIEFING_DATE"}

HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E79")


# ── Folder structure ──────────────────────────────────────────────────────────

def create_batch_folder(date_from: str, date_to: str, batch_type: str,
                        root_dir: str = None) -> str:
    """
    Create and return the root batch folder.
    root_dir defaults to data/etenders.gov.za (the only mode this deployment runs).
    """
    if root_dir is None:
        root_dir = os.path.join("data", "etenders.gov.za")

    start = datetime.strptime(date_from, "%Y-%m-%d")
    end   = datetime.strptime(date_to,   "%Y-%m-%d")

    if start.month == end.month:
        label = f"({batch_type}) {start.day}-{end.day} {end.strftime('%B %Y')}"
    else:
        label = (f"({batch_type}) {start.strftime('%d %b')}"
                 f"-{end.strftime('%d %b %Y')}")

    import shutil
    root = os.path.join(root_dir, label)
    if os.path.exists(root):
        try:
            shutil.rmtree(root)
            logging.info(f"Existing batch folder removed for re-scrape: {root}")
        except PermissionError:
            logging.warning(f"Could not delete existing batch folder (files open?); will overwrite in place: {root}")
    os.makedirs(os.path.join(root, "batches"),          exist_ok=True)
    os.makedirs(os.path.join(root, "end product"),      exist_ok=True)
    os.makedirs(os.path.join(root, "Display Equation"), exist_ok=True)
    logging.info(f"Batch folder ready: {root}")
    return root


# ── Excel helpers ─────────────────────────────────────────────────────────────

def write_tender_rows(ws, df: pd.DataFrame, columns: list = None) -> None:
    """Write headers + data rows to a Tender Data worksheet."""
    if columns is None:
        columns = TENDER_COLUMNS
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    ordered = df.reindex(columns=columns)
    for row_idx, row in enumerate(ordered.itertuples(index=False), 2):
        for col_idx, (col_name, value) in enumerate(zip(columns, row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in DATE_COLUMNS and isinstance(value, str) and value:
                try:
                    cell.value = datetime.strptime(value, "%Y/%m/%d")
                    cell.number_format = "YYYY/MM/DD"
                except ValueError:
                    pass
            if col_name == "LINK" and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"


def _write_final_sheet(ws, batch_type: str, report_date: datetime,
                       no_etenders: bool = False) -> None:
    """Write the Final worksheet with the same formulas as the existing files.

    When no_etenders=True (All-but-eTenders mode), the "eTenders (All)" row is
    forced to 0 since no eTenders portal scrape occurred, and a labelled TOTAL
    row is written to reflect the sum of the tracked entities.
    """
    for col_idx, heading in enumerate(
        ["Source", "Number of New Tenders", "ICT Tenders", "RFQs"], 1
    ):
        cell = ws.cell(row=1, column=col_idx, value=heading)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    ws["A2"] = "Report Date"
    ws["B2"] = f"({batch_type}) {report_date.strftime('%d/%m/%Y')}"

    ws["A3"] = "eTenders (All)"
    if no_etenders:
        ws["B3"] = 0
        ws["C3"] = 0
        ws["D3"] = 0
    else:
        ws["B3"] = "=COUNTA('Tender Data'!J:J)-1"
        ws["C3"] = (
            "=COUNTIF('Tender Data'!M:M,\"Information service activities\")"
            "+COUNTIF('Tender Data'!M:M,\"Information and communication\")"
        )
        ws["D3"] = "=COUNTIF('Tender Data'!G:G,\"Request for Quotation\")"

    for offset, (label, pattern) in enumerate(SOURCES[1:]):
        r = 4 + offset
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2,
                value=f"=COUNTIF('Tender Data'!J:J,\"{pattern}\")")
        ws.cell(row=r, column=3, value=(
            f"=COUNTIFS('Tender Data'!M:M,\"Information service activities\","
            f"'Tender Data'!J:J,\"{pattern}\")"
            f"+COUNTIFS('Tender Data'!M:M,\"Information and communication\","
            f"'Tender Data'!J:J,\"{pattern}\")"
        ))
        ws.cell(row=r, column=4, value=(
            f"=COUNTIFS('Tender Data'!J:J,\"{pattern}\","
            f"'Tender Data'!G:G,\"Request for Quotation\")"
        ))

    total_row = 3 + len(SOURCES)  # first row after last entity
    ws.cell(row=total_row, column=1, value="TOTAL WATCHLIST")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row - 1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row - 1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row - 1})")
    for c in (2, 3, 4):
        ws.cell(row=total_row, column=c).font = Font(bold=True)


# ── Per-day file ──────────────────────────────────────────────────────────────

def save_daily_file(tender_data: list, day: str, batch_folder: str) -> str:
    """
    Save one day's tender data to {batch_folder}/batches/tenders_YYYY_MM_DD.xlsx.
    If tender_data is empty the filename gets a ' (No Tenders)' suffix so the
    day is still accounted for in the batches folder.
    Returns the saved file path.
    """
    date_str  = day.replace("-", "_")
    suffix    = " (No Tenders)" if not tender_data else ""
    filepath  = os.path.join(batch_folder, "batches", f"tenders_{date_str}{suffix}.xlsx")

    # Assign record IDs (first scraped = highest number)
    records = []
    total = len(tender_data)
    for idx, tender in enumerate(tender_data, 1):
        t = tender.copy()
        t["RECORD_ID"] = total - idx + 1
        records.append(t)

    df = pd.DataFrame(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tender Data"
    write_tender_rows(ws, df, BATCH_COLUMNS)
    wb.save(filepath)

    logging.info(f"Daily file saved: {filepath} ({total} tenders)")
    return filepath


# ── End product file ──────────────────────────────────────────────────────────

def create_end_product(df: pd.DataFrame, date_from: str, date_to: str,
                       batch_type: str, report_date: datetime,
                       batch_folder: str,
                       raw_df: pd.DataFrame = None,
                       no_etenders: bool = False) -> str:
    """
    Create the end product Excel file inside {batch_folder}/end product/.
    Sheets: Tender Data (deduplicated), Final (formulas), Raw Data (all scraped, optional).
    Returns the saved file path.
    """
    start = datetime.strptime(date_from, "%Y-%m-%d")
    end   = datetime.strptime(date_to,   "%Y-%m-%d")

    if start.month == end.month:
        label = f"{start.day} - {end.day} {end.strftime('%B')}"
    else:
        label = f"{start.strftime('%d %b')} - {end.strftime('%d %b')}"

    filename = f"RFQ_and_ICT_Checker_({label}).xlsx"
    filepath = os.path.join(batch_folder, "end product", filename)

    # Assign record IDs across the full combined dataset
    records = []
    total = len(df)
    for idx, row in enumerate(df.itertuples(index=False), 1):
        t = {col: getattr(row, col, None) for col in TENDER_COLUMNS if col != "RECORD_ID"}
        t["RECORD_ID"] = total - idx + 1
        records.append(t)

    df_out = pd.DataFrame(records)

    wb = Workbook()
    wb.remove(wb.active)
    write_tender_rows(wb.create_sheet("Tender Data"), df_out, BATCH_COLUMNS)
    _write_final_sheet(wb.create_sheet("Final"), batch_type, report_date,
                       no_etenders=no_etenders)

    if raw_df is not None and not raw_df.empty:
        raw_records = []
        raw_total = len(raw_df)
        for idx, row in enumerate(raw_df.itertuples(index=False), 1):
            t = {col: getattr(row, col, None) for col in TENDER_COLUMNS if col != "RECORD_ID"}
            t["RECORD_ID"] = raw_total - idx + 1
            raw_records.append(t)
        write_tender_rows(wb.create_sheet("Raw Data"), pd.DataFrame(raw_records), BATCH_COLUMNS)
        logging.info(f"Raw Data sheet written: {raw_total} rows ({raw_total - total} duplicates)")

    wb.save(filepath)
    logging.info(f"End product created: {filepath}")
    return filepath


# ── Deduplication ────────────────────────────────────────────────────────────

def _normalize_tender_id(tid: str) -> str:
    """Normalize a tender ID for cross-source deduplication comparison.
    Handles cases like eTenders appending 'ERP No:424011' to SITA IDs,
    and SITA using underscores/hyphens where eTenders uses slashes.
    """
    tid = re.sub(r'\s+ERP\s+No[:.]\s*\d+', '', tid, flags=re.IGNORECASE).strip()
    tid = re.sub(r'[-_]', '/', tid)
    return tid.upper()


def filter_to_watchlist(tenders: list) -> list:
    """Keep only tenders whose DEPARTMENT matches a SOURCES pattern.

    Used by eTenders mode and Full Batch mode to drop tenders from organs of
    state that are not on the watchlist. The removed rows can still be kept
    upstream (e.g. in the Raw Data sheet) for spot-checking.
    """
    patterns = [p for _, p in SOURCES if p is not None]
    def _matches(t):
        dept = str(t.get("DEPARTMENT") or "")
        return any(_xl_match(dept, p) for p in patterns)
    return [t for t in tenders if _matches(t)]


def merge_and_flag_duplicates(tenders: list, prefer_etenders: bool = True) -> list:
    """Cross-source dedupe with a DUPLICATED flag column.

    Groups tenders by normalized TENDER_ID. When a group has more than one
    entry (i.e. the same tender was captured from both the eTenders portal AND
    an organ's own website), keep exactly one row and set DUPLICATED=1. Single-
    occurrence groups get DUPLICATED=0.

    If prefer_etenders is True, the eTenders portal version is kept when there
    is a conflict (identified via the internal '_from_etenders' marker). This
    is usually preferable because the portal populates every field.

    Tenders with a blank TENDER_ID are always kept (never treated as duplicates
    of each other), because there's no reliable key to compare on.
    """
    groups: dict = {}
    blanks: list = []
    for t in tenders:
        tid = str(t.get("TENDER_ID") or "").strip()
        if not tid:
            blanks.append(t)
            continue
        key = _normalize_tender_id(tid)
        groups.setdefault(key, []).append(t)

    merged: list = []
    for _, ts in groups.items():
        if len(ts) > 1:
            if prefer_etenders:
                et_versions = [t for t in ts if t.get("_from_etenders")]
                chosen = et_versions[0] if et_versions else ts[0]
            else:
                chosen = ts[0]
            chosen["DUPLICATED"] = 1
        else:
            chosen = ts[0]
            chosen["DUPLICATED"] = 0
        chosen.pop("_from_etenders", None)
        merged.append(chosen)

    for t in blanks:
        t["DUPLICATED"] = 0
        t.pop("_from_etenders", None)
        merged.append(t)

    return merged


def deduplicate_tenders(tenders: list) -> list:
    """
    Remove duplicate tender entries by TENDER_ID (first occurrence wins).
    Uses normalized IDs so cross-source duplicates (e.g. SITA on both
    eTenders.gov.za and sita.co.za) are caught even when IDs differ slightly.
    Records with a blank/None TENDER_ID are always kept as-is.
    """
    seen: set = set()
    out: list = []
    dupes = 0
    for t in tenders:
        tid  = str(t.get("TENDER_ID") or "").strip()
        norm = _normalize_tender_id(tid) if tid else ""
        if norm and norm in seen:
            dupes += 1
            logging.info(f"Duplicate tender skipped: {tid}")
            continue
        out.append(t)
        if norm:
            seen.add(norm)
    if dupes:
        logging.info(f"Deduplication removed {dupes} duplicate tender(s)")
    return out


# ── Counting ──────────────────────────────────────────────────────────────────

def _xl_match(value: str, pattern: str) -> bool:
    if not isinstance(value, str):
        return False
    regex = re.escape(pattern).replace(r"\*", ".*")
    return bool(re.fullmatch(regex, value, re.IGNORECASE))


def calculate_counts(df: pd.DataFrame) -> dict:
    dept = df["DEPARTMENT"].fillna("")
    cat  = df["CATEGORY"].fillna("")
    typ  = df["TENDER_TYPE"].fillna("")

    is_ict = cat.str.lower().isin([c.lower() for c in ICT_CATEGORIES])
    is_rfq = typ.str.lower() == RFQ_TYPE.lower()

    result = {}
    for label, pattern in SOURCES:
        if pattern is None:
            mask = pd.Series([True] * len(df), index=df.index)
        else:
            mask = dept.apply(lambda v, p=pattern: _xl_match(v, p))
        result[label] = {
            "NNT": int(mask.sum()),
            "ICT": int((mask & is_ict).sum()),
            "RFQ": int((mask & is_rfq).sum()),
        }
    return result


# ── Equation file ─────────────────────────────────────────────────────────────

MAX_EQUATION_BATCHES = 6


def update_equation_file(counts: dict, batch_type: str, report_date: datetime,
                         batch_folder: str, no_etenders: bool = False) -> None:
    """
    Insert a new NNT/ICT/RFQ column block at column B (newest-first), shift
    older batches right, trim to MAX_EQUATION_BATCHES, then copy to the
    Display Equation folder inside batch_folder.

    When no_etenders=True (All-but-eTenders mode), the "eTenders (All)" cell for
    this batch is forced to 0. A "TOTAL" row is always written below the last
    source, summing every source above it.
    """
    import shutil

    if os.path.exists(EQUATION_FILE):
        wb = load_workbook(EQUATION_FILE)
        ws = wb["Sheet1"]
    else:
        parent = os.path.dirname(EQUATION_FILE)
        if parent:
            os.makedirs(parent, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    batch_label = f"({batch_type}) {report_date.strftime('%d/%m/%y')}"

    # Remove any existing column for this batch label (handles re-runs)
    col = 2
    while ws.cell(row=1, column=col).value is not None:
        if ws.cell(row=1, column=col).value == batch_label:
            ws.delete_cols(col, 3)
            break
        col += 3

    # Insert 3 blank columns at position B, pushing existing data right
    ws.insert_cols(2, 3)

    # Write new batch in columns B, C, D (indices 2, 3, 4)
    ws.cell(row=1, column=2, value=batch_label)
    ws.cell(row=2, column=2, value="NNT")
    ws.cell(row=2, column=3, value="ICT")
    ws.cell(row=2, column=4, value="RFQ")

    for row_offset, (source, _) in enumerate(SOURCES):
        data = counts.get(source, {"NNT": 0, "ICT": 0, "RFQ": 0})
        if no_etenders and source == "eTenders (All)":
            data = {"NNT": 0, "ICT": 0, "RFQ": 0}
        ws.cell(row=3 + row_offset, column=1, value=source)
        ws.cell(row=3 + row_offset, column=2, value=data["NNT"])
        ws.cell(row=3 + row_offset, column=3, value=data["ICT"])
        ws.cell(row=3 + row_offset, column=4, value=data["RFQ"])

    # TOTAL row below the last source — sums NNT/ICT/RFQ for the tracked
    # entities (rows 4..last_source, skipping the eTenders (All) row on row 3).
    total_row = 3 + len(SOURCES)  # first row after last source
    ws.cell(row=total_row, column=1, value="TOTAL WATCHLIST").font = Font(bold=True)
    for c in (2, 3, 4):
        letter = ws.cell(row=total_row, column=c).column_letter
        ws.cell(row=total_row, column=c,
                value=f"=SUM({letter}4:{letter}{total_row - 1})").font = Font(bold=True)

    # Count batches and trim oldest if over the limit
    num_batches = 0
    col = 2
    while ws.cell(row=1, column=col).value is not None:
        num_batches += 1
        col += 3

    while num_batches > MAX_EQUATION_BATCHES:
        oldest_col = 2 + (num_batches - 1) * 3
        ws.delete_cols(oldest_col, 3)
        num_batches -= 1

    # Apply thin borders to the full data range (headers + sources + TOTAL row)
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    total_rows = 2 + len(SOURCES) + 1  # +1 for TOTAL row
    total_cols = 1 + num_batches * 3
    for r in range(1, total_rows + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = _border

    wb.save(EQUATION_FILE)
    logging.info(f"Equation file updated: {batch_label} (no_etenders={no_etenders})")

    # Copy updated equation file to Display Equation folder
    dest = os.path.join(batch_folder, "Display Equation",
                        os.path.basename(EQUATION_FILE))
    shutil.copy2(EQUATION_FILE, dest)
    logging.info(f"Display Equation copy saved: {dest}")

    # All-but-eTenders mode: inject a static Website (Yes/No) column into the
    # batch-folder copy only. Master equation file stays untouched.
    if no_etenders:
        _inject_website_column(dest, len(SOURCES))


def _inject_website_column(path: str, num_sources: int) -> None:
    """Insert a 'Website' column at position B into the equation copy at path.
    Populates Yes/No per source from WATCHLIST_HAS_SCRAPER."""
    wb = load_workbook(path)
    ws = wb["Sheet1"]
    ws.insert_cols(2)

    header = ws.cell(row=2, column=2, value="Website")
    header.font = Font(bold=True, color="FFFFFF")
    header.fill = HDR_FILL
    header.alignment = Alignment(horizontal='center')

    _yes_fill = PatternFill('solid', fgColor='C6EFCE')  # pale green
    _no_fill  = PatternFill('solid', fgColor='FFC7CE')  # pale red
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    for row_offset, (source, _) in enumerate(SOURCES):
        yes_no = WATCHLIST_HAS_SCRAPER.get(source, "-")
        cell = ws.cell(row=3 + row_offset, column=2, value=yes_no)
        cell.alignment = Alignment(horizontal='center')
        cell.border = _border
        if yes_no == "Yes":
            cell.fill = _yes_fill
        elif yes_no == "No":
            cell.fill = _no_fill

    # Header cell + TOTAL row's Website cell get borders too
    header.border = _border
    total_row = 3 + num_sources
    ws.cell(row=total_row, column=2).border = _border

    # Set column B width
    ws.column_dimensions[ws.cell(row=2, column=2).column_letter].width = 10

    wb.save(path)
    logging.info(f"Website column injected into: {path}")


# ── Power BI export ───────────────────────────────────────────────────────────

PBI_COLUMNS = ["BATCH_LABEL", "BATCH_ORDER", "BATCH_TYPE", "BATCH_START_DATE", "BATCH_SORT_KEY", "REPORT_DATE", "SOURCE", "SOURCE_SORT", "NNT", "ICT", "RFQ"]

# Sort key per SOURCE: 0 = eTenders (All) pinned first, 1+ = alphabetical order of the rest
_SOURCE_SORT = {label: (0 if label == "eTenders (All)" else idx) for idx, (label, _) in enumerate(SOURCES)}


def update_power_bi_export(batch_folder: str, date_from: str, date_to: str, batch_type: str) -> None:
    """
    Append (or replace) this batch's rows in the running Power BI data file.
    One row per SOURCE per day — reads each daily batch file from batch_folder/batches/.
    Existing rows for this batch_label are replaced on re-runs.
    """
    start = datetime.strptime(date_from, "%Y-%m-%d")
    end   = datetime.strptime(date_to,   "%Y-%m-%d")

    if start.month == end.month:
        batch_label = f"({batch_type}) {start.day}-{end.day} {end.strftime('%b %y')}"
    else:
        batch_label = f"({batch_type}) {start.strftime('%d %b')}-{end.strftime('%d %b %y')}"

    # Read eTenders (All) totals from the Display Equation file — these are the
    # authoritative totals already calculated by the scraper pipeline, covering
    # all ICT/RFQ categories correctly. Do not recalculate from raw batch data.
    eq_file = os.path.join(batch_folder, "Display Equation",
                           os.path.basename(EQUATION_FILE))
    etenders_all_total = {"NNT": 0, "ICT": 0, "RFQ": 0}
    if os.path.exists(eq_file):
        try:
            eq_wb = load_workbook(eq_file, data_only=True)
            eq_ws = eq_wb["Sheet1"]
            # Row 3 = eTenders (All), cols B/C/D = NNT/ICT/RFQ of newest batch
            nnt = eq_ws.cell(3, 2).value
            ict = eq_ws.cell(3, 3).value
            rfq = eq_ws.cell(3, 4).value
            etenders_all_total = {
                "NNT": int(nnt) if isinstance(nnt, (int, float)) else 0,
                "ICT": int(ict) if isinstance(ict, (int, float)) else 0,
                "RFQ": int(rfq) if isinstance(rfq, (int, float)) else 0,
            }
            logging.info(f"Power BI: eTenders (All) totals from Display Equation — "
                         f"NNT={etenders_all_total['NNT']} ICT={etenders_all_total['ICT']} "
                         f"RFQ={etenders_all_total['RFQ']}")
        except Exception as e:
            logging.warning(f"Could not read Display Equation for eTenders (All): {e}")

    # Build new rows — one per source per day
    new_rows = []
    batches_dir = os.path.join(batch_folder, "batches")

    current = start
    while current <= end:
        day_str  = current.strftime("%Y_%m_%d")
        day_file = os.path.join(batches_dir, f"tenders_{day_str}.xlsx")

        if os.path.exists(day_file):
            day_df     = pd.read_excel(day_file)
            day_counts = calculate_counts(day_df)
        else:
            day_counts = {source: {"NNT": 0, "ICT": 0, "RFQ": 0} for source, _ in SOURCES}

        # Override eTenders (All) with the authoritative Display Equation total
        # (only on the report date — last day of batch — to avoid splitting across days)
        day_counts["eTenders (All)"] = etenders_all_total if current == end else {"NNT": 0, "ICT": 0, "RFQ": 0}

        for source, _ in SOURCES:
            data = day_counts.get(source, {"NNT": 0, "ICT": 0, "RFQ": 0})
            new_rows.append({
                "BATCH_LABEL":      batch_label,
                "BATCH_TYPE":       batch_type,
                "BATCH_START_DATE": start,
                "BATCH_SORT_KEY":   -start.toordinal(),
                "REPORT_DATE":      current,
                "SOURCE":           source,
                "SOURCE_SORT":      _SOURCE_SORT.get(source, 999),
                "NNT":              data["NNT"],
                "ICT":              data["ICT"],
                "RFQ":              data["RFQ"],
            })

        current += timedelta(days=1)

    new_df = pd.DataFrame(new_rows, columns=[c for c in PBI_COLUMNS if c != "BATCH_ORDER"])

    # Load existing data and drop any prior rows for this batch (idempotent)
    if os.path.exists(POWER_BI_FILE):
        try:
            existing = pd.read_excel(POWER_BI_FILE)
            existing = existing[existing["BATCH_LABEL"] != batch_label]
            combined = pd.concat([existing, new_df], ignore_index=True)
        except Exception:
            combined = new_df
    else:
        combined = new_df

    # Sort newest batch first — keep ALL batches, no rolling trim
    combined["REPORT_DATE"] = pd.to_datetime(combined["REPORT_DATE"])
    combined["BATCH_START_DATE"] = pd.to_datetime(combined["BATCH_START_DATE"])
    combined.sort_values("REPORT_DATE", ascending=False, inplace=True)
    combined.reset_index(drop=True, inplace=True)

    # Assign BATCH_ORDER: 1 = newest batch, incrementing for older batches
    unique_batches = (
        combined[["BATCH_LABEL", "BATCH_START_DATE"]]
        .drop_duplicates("BATCH_LABEL")
        .sort_values("BATCH_START_DATE", ascending=False)
        .reset_index(drop=True)
    )
    unique_batches["BATCH_ORDER"] = unique_batches.index + 1
    order_map = dict(zip(unique_batches["BATCH_LABEL"], unique_batches["BATCH_ORDER"]))
    combined["BATCH_ORDER"] = combined["BATCH_LABEL"].map(order_map)
    combined = combined[PBI_COLUMNS]

    # Write with styled header and named Excel Table for Power BI compatibility
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(POWER_BI_FILE), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tender Data"

    for col_idx, col_name in enumerate(PBI_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    for row_idx, row in enumerate(combined.itertuples(index=False), 2):
        for col_idx, col_name in enumerate(PBI_COLUMNS, 1):
            value = getattr(row, col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in ("REPORT_DATE", "BATCH_START_DATE") and hasattr(value, "strftime"):
                cell.number_format = "YYYY/MM/DD"

    last_col = get_column_letter(len(PBI_COLUMNS))
    last_row = len(combined) + 1
    tbl = Table(displayName="TenderData", ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    wb.save(POWER_BI_FILE)
    logging.info(f"Power BI export updated: {batch_label} ({len(new_rows)} rows) -> {POWER_BI_FILE}")


# ── Master tender ledger (cumulative, never trimmed) ──────────────────────────

MASTER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "master_tenders.xlsx")
_MASTER_OUTPUT_COLS = [c for c in TENDER_COLUMNS if c != "RECORD_ID"]


def update_master_tenders(batch_folder: str) -> None:
    """
    Append every tender from this batch's daily files into the master ledger.
    Deduplicates by TENDER_ID so re-runs are safe. New data takes priority.
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    batches_dir = os.path.join(batch_folder, "batches")
    new_frames  = []
    for fname in sorted(os.listdir(batches_dir)):
        if not fname.endswith(".xlsx") or "No Tenders" in fname:
            continue
        try:
            df = pd.read_excel(os.path.join(batches_dir, fname), dtype=str)
            new_frames.append(df)
        except Exception as e:
            logging.warning(f"Master update — could not read {fname}: {e}")

    if not new_frames:
        logging.info("Master update — no new tenders in batch")
        return

    new_df = pd.concat(new_frames, ignore_index=True)
    new_df.columns = [c.strip() for c in new_df.columns]
    new_df = new_df[[c for c in _MASTER_OUTPUT_COLS if c in new_df.columns]]
    for col in _MASTER_OUTPUT_COLS:
        if col not in new_df.columns:
            new_df[col] = "AUTOMATIC" if col == "INGESTION_METHOD" else None
    new_df = new_df[_MASTER_OUTPUT_COLS]

    # Load existing master
    if os.path.exists(MASTER_FILE):
        try:
            existing = pd.read_excel(MASTER_FILE, dtype=str)
            from master_schema import read_template_schema
            existing = read_template_schema(existing)
            existing = existing[[c for c in _MASTER_OUTPUT_COLS if c in existing.columns]]
            for col in _MASTER_OUTPUT_COLS:
                if col not in existing.columns:
                    existing[col] = None
            existing = existing[_MASTER_OUTPUT_COLS]
        except Exception:
            existing = pd.DataFrame(columns=_MASTER_OUTPUT_COLS)
    else:
        existing = pd.DataFrame(columns=_MASTER_OUTPUT_COLS)

    # New data first so it wins deduplication
    combined = pd.concat([new_df, existing], ignore_index=True)

    seen  = set()
    keep  = []
    dupes = 0
    for idx, row in combined.iterrows():
        tid  = str(row.get("TENDER_ID") or "").strip()
        norm = _normalize_tender_id(tid) if tid else ""
        if norm and norm in seen:
            dupes += 1
            continue
        keep.append(idx)
        if norm:
            seen.add(norm)
    combined = combined.loc[keep].reset_index(drop=True)
    logging.info(f"Master update — {len(new_df)} new, {dupes} duplicates skipped, {len(combined)} total")

    # Sort newest first
    combined["REPORT_DATE"] = pd.to_datetime(combined["REPORT_DATE"], errors="coerce")
    combined.sort_values("REPORT_DATE", ascending=False, inplace=True)
    combined.reset_index(drop=True, inplace=True)

    combined.insert(1, "RECORD_ID", range(1, len(combined) + 1))

    from master_schema import apply_template_schema, TEMPLATE_COLUMN_ORDER
    combined = apply_template_schema(combined)
    all_cols = TEMPLATE_COLUMN_ORDER

    from openpyxl.styles import Alignment, Border, Side

    MASTER_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    MASTER_HDR_FILL = PatternFill("solid", fgColor="305496")
    MASTER_ROW_FILL_ODD  = PatternFill("solid", fgColor="DDEBF7")
    MASTER_ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="9DC3E6"),
        right=Side(style="thin", color="9DC3E6"),
    )

    os.makedirs(os.path.dirname(MASTER_FILE), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "TENDER REPORT"

    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = MASTER_HDR_FONT
        cell.fill = MASTER_HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    _row_values = combined[all_cols].to_numpy()
    for row_idx in range(2, len(_row_values) + 2):
        row_fill = MASTER_ROW_FILL_ODD if row_idx % 2 == 0 else MASTER_ROW_FILL_EVEN
        row_arr = _row_values[row_idx - 2]
        for col_idx, col_name in enumerate(all_cols, 1):
            value = row_arr[col_idx - 1]
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            if col_name in DATE_COLUMNS and value and not pd.isna(value):
                try:
                    if not isinstance(value, datetime):
                        value = pd.to_datetime(value)
                    cell.value = value.to_pydatetime()
                    cell.number_format = "YYYY/MM/DD"
                except Exception:
                    pass

    # Auto-fit column widths
    for col_idx, col_name in enumerate(all_cols, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(combined[col_name].iloc[i])) for i in range(min(50, len(combined)))]
        ) if len(combined) > 0 else len(col_name)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    last_col = get_column_letter(len(all_cols))
    last_row = len(combined) + 1
    tbl = Table(displayName="MasterTenders", ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    wb.save(MASTER_FILE)
    logging.info(f"Master tender file saved: {MASTER_FILE} ({len(combined)} total rows)")
