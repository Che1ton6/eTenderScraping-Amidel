#!/usr/bin/env python3
"""Build the manual-scrape template workbook for Thandolwethu.

One tab per problematic watchlist source. Each tab lists the fields we
normally capture with a status flag indicating whether the scraper is
succeeding on that field for that source. She fills in the "Resolvable?"
and "Notes" columns as she works through them.

Output: Sales\Sales Auto Hub\Scraping and Reports\Manual_Scrape_Template.xlsx
"""
import io, sys, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
BATCH = os.path.join(HERE, 'data', 'All_Tenders', '(M) 9-12 July 2026')
END_PRODUCT = os.path.join(BATCH, 'end product', 'RFQ_and_ICT_Checker_(9 - 12 July).xlsx')

USER = os.path.expanduser('~')
DEST_DIR = os.path.join(USER, 'OneDrive - Amidel (Pty) Ltd', 'Documents', 'Sales',
                        'Sales Auto Hub', 'Scraping and Reports')
OUT = os.path.join(DEST_DIR, 'Manual_Scrape_Template.xlsx')
OLD_DOWNLOAD = os.path.join(USER, 'Downloads', 'Manual_Scrape_Template.xlsx')

ENTITIES = [
    ('Umzimvubu LM',          'UMZIMVUBU.GOV.ZA',        'Umzimvubu LM',
        'https://www.umzimvubu.gov.za'),
    ('Mnquma LM',             'MNQUMA.GOV.ZA',           'Mnquma LM',
        'https://www.mnquma.gov.za'),
    ('Matatiele LM',          'MATATIELE.GOV.ZA',        'Matatiele LM',
        'https://www.matatiele.gov.za'),
    ('Nelson Mandela Bay MM', 'NELSONMANDELABAY.GOV.ZA', 'Nelson Mandela Bay',
        'https://www.nelsonmandelabay.gov.za'),
    ('Great Kei LM',          'GREATKEILM.GOV.ZA',       'Great Kei LM',
        'https://www.greatkeilm.gov.za'),
    ('Ntabankulu LM',         'NTABANKULU.GOV.ZA',       'Ntabankulu LM',
        'https://www.ntabankulu.gov.za'),
    ('Buffalo City MM',       'BUFFALOCITY.GOV.ZA',      'Buffalo City MM',
        'https://www.buffalocity.gov.za'),
    ('GDoH',                  None,                      'GDoH',
        'http://professionaljobcentre.gpg.gov.za'),
    ('Winnie Mandela LM',     'WINNIEMMLM.GOV.ZA',       'Winnie Mandela LM',
        'https://www.winniemmlm.gov.za'),
    ('Amahlathi LM',          'AMAHLATHI.GOV.ZA',        'Amahlathi LM',
        'https://www.amahlathi.gov.za'),
    ('GPL',                   'GPL.GOV.ZA',              'GPL',
        'https://www.gpl.gov.za'),
]

FIELDS = [
    'TENDER_ID', 'TENDER_DESCRIPTION', 'DEPARTMENT', 'PROVINCE',
    'PUBLICATION_DATE', 'CLOSING_DATE', 'CLOSING_TIME',
    'TENDER_TYPE', 'CATEGORY', 'LINK',
]

HDR_FONT   = Font(bold=True, color='FFFFFF')
HDR_FILL   = PatternFill('solid', fgColor='1C3880')
TRACK_FILL = PatternFill('solid', fgColor='F5A000')
MISS_FILL  = PatternFill('solid', fgColor='FBD5B5')  # amber for issues
OK_FILL    = PatternFill('solid', fgColor='C6E0B4')  # green for OK
_thin      = Side(style='thin', color='C0C0C0')
BORDER     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _valid_date(val):
    try:
        pd.to_datetime(str(val).replace('/', '-'))
        return True
    except Exception:
        return False


def _field_status(series, field):
    """Return (status, detail) for a field across all rows of one entity."""
    n = len(series)
    if n == 0:
        return ('No data scraped', f'0 rows scraped this batch')
    blanks = series.fillna('').astype(str).str.strip().eq('').sum()
    if field in ('PUBLICATION_DATE', 'CLOSING_DATE'):
        non_blank = series.fillna('').astype(str).str.strip().ne('')
        invalid = (non_blank & ~series.apply(_valid_date)).sum()
        if blanks == n:
            return ('Missing', f'{n}/{n} rows blank')
        if invalid + blanks == n:
            return ('Invalid format', f'{invalid}/{n} rows have unparseable dates, {blanks}/{n} blank')
        if invalid + blanks > 0:
            return ('Partial', f'{blanks}/{n} blank, {invalid}/{n} invalid')
        return ('Captured', f'{n}/{n} rows have valid dates')
    if blanks == n:
        return ('Missing', f'{n}/{n} rows blank')
    if blanks > 0:
        return ('Partial', f'{blanks}/{n} blank')
    return ('Captured', f'{n}/{n} rows populated')


def _for_entity(raw_df, domain):
    if domain is None:
        dept = raw_df['DEPARTMENT'].fillna('').astype(str).str.lower()
        return raw_df[dept.str.contains('gauteng') & dept.str.contains('health')]
    src = raw_df['TENDER_SOURCE'].fillna('').astype(str).str.upper()
    return raw_df[src == domain]


def _apply_status_fill(cell, status):
    if status == 'Captured':
        cell.fill = OK_FILL
    else:
        cell.fill = MISS_FILL


def main():
    print(f'Reading {END_PRODUCT}')
    raw = pd.read_excel(END_PRODUCT, sheet_name='Raw Data')
    print(f'Raw rows: {len(raw)}')

    wb = Workbook()
    wb.remove(wb.active)

    # ── Index ────────────────────────────────────────────────────────────────
    idx = wb.create_sheet('Index', 0)
    idx['A1'] = 'Manual Scrape Template'
    idx['A1'].font = Font(bold=True, size=14, color='1C3880')
    idx['A2'] = f'Batch reference: (M) 9-12 July 2026  |  Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    idx['A2'].font = Font(italic=True, color='555555')
    idx['A4'] = ('Each tab is one watchlist entity. It lists the tender fields we normally capture and shows '
                 'whether our scraper is succeeding on that field for that site. Where the status is not '
                 '"Captured", please investigate the source website and mark whether the field can be found '
                 'there. Use the Resolvable column and Notes to help us prioritise scraper fixes.')
    idx['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    idx.merge_cells('A4:F4')
    idx.row_dimensions[4].height = 60

    for c, name in enumerate(['Entity', 'Website', 'Rows scraped this batch', 'Fields OK', 'Fields with issues'], 1):
        cell = idx.cell(row=6, column=c, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    # ── Per-entity tabs ──────────────────────────────────────────────────────
    for i, (label, domain, tab, url) in enumerate(ENTITIES):
        sub = _for_entity(raw, domain)
        n = len(sub)

        ws = wb.create_sheet(tab)
        ws['A1'] = label
        ws['A1'].font = Font(bold=True, size=14, color='1C3880')
        ws['A2'] = f'Website: {url}'
        ws['A2'].font = Font(color='1C3880', underline='single')
        ws['A2'].hyperlink = url
        ws['A3'] = f'Rows this batch: {n}'
        ws['A3'].font = Font(italic=True, color='555555')

        headers = ['Field', 'Scraper Status', 'Detail', 'Resolvable?', 'Notes']
        for c, name in enumerate(headers, 1):
            cell = ws.cell(row=5, column=c, value=name)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL if c <= 3 else TRACK_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')

        ok_count = 0
        issue_count = 0
        for r, field in enumerate(FIELDS, 6):
            if n == 0:
                status, detail = ('No data scraped', '0 rows scraped this batch')
            else:
                status, detail = _field_status(sub[field], field)
            if status == 'Captured':
                ok_count += 1
            else:
                issue_count += 1
            ws.cell(row=r, column=1, value=field).border = BORDER
            status_cell = ws.cell(row=r, column=2, value=status)
            status_cell.border = BORDER
            status_cell.alignment = Alignment(horizontal='center')
            status_cell.font = Font(bold=True)
            _apply_status_fill(status_cell, status)
            ws.cell(row=r, column=3, value=detail).border = BORDER
            ws.cell(row=r, column=4, value='').border = BORDER
            ws.cell(row=r, column=5, value='').border = BORDER

        # Dropdowns for the last two columns
        last_row = 5 + len(FIELDS)
        dv = DataValidation(type='list', formula1='"Yes,No,Unknown"', allow_blank=True)
        dv.add(f'D6:D{last_row}')
        ws.add_data_validation(dv)

        # Column widths
        for col_letter, w in zip('ABCDE', [22, 22, 55, 15, 42]):
            ws.column_dimensions[col_letter].width = w
        ws.row_dimensions[5].height = 28
        ws.freeze_panes = 'A6'

        # Index entry
        r = 7 + i
        idx.cell(row=r, column=1, value=label).border = BORDER
        idx.cell(row=r, column=2, value=domain if domain else '(varies)').border = BORDER
        idx.cell(row=r, column=3, value=n).border = BORDER
        idx.cell(row=r, column=4, value=ok_count).border = BORDER
        idx.cell(row=r, column=5, value=issue_count).border = BORDER
        for c in range(3, 6):
            idx.cell(row=r, column=c).alignment = Alignment(horizontal='center')

        print(f'  {label:<24} rows={n:<3} OK={ok_count}  issues={issue_count}')

    for col_letter, w in zip('ABCDE', [26, 30, 24, 14, 20]):
        idx.column_dimensions[col_letter].width = w

    # ── Save + relocate ──────────────────────────────────────────────────────
    os.makedirs(DEST_DIR, exist_ok=True)
    wb.save(OUT)
    print(f'\nWROTE: {OUT}')

    if os.path.exists(OLD_DOWNLOAD):
        try:
            os.remove(OLD_DOWNLOAD)
            print(f'Removed old copy: {OLD_DOWNLOAD}')
        except PermissionError:
            print(f'WARNING: could not remove old copy (file is open): {OLD_DOWNLOAD}')


if __name__ == '__main__':
    main()
