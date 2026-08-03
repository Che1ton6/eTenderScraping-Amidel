"""One-time script: rebuilds eTender_PowerBI_Data.xlsx in daily-breakdown format from all batch folders."""
import os
import re
import sys
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from BatchProcessor import SOURCES, calculate_counts, _SOURCE_SORT, EQUATION_FILE


def read_etenders_all_from_equation(folder_path) -> dict:
    """Read eTenders (All) NNT/ICT/RFQ from the Display Equation file in this batch folder."""
    eq_file = os.path.join(folder_path, "Display Equation", os.path.basename(EQUATION_FILE))
    if not os.path.exists(eq_file):
        return None
    try:
        wb = load_workbook(eq_file, data_only=True)
        ws = wb["Sheet1"]
        nnt = ws.cell(3, 2).value
        ict = ws.cell(3, 3).value
        rfq = ws.cell(3, 4).value
        return {
            "NNT": int(nnt) if isinstance(nnt, (int, float)) else 0,
            "ICT": int(ict) if isinstance(ict, (int, float)) else 0,
            "RFQ": int(rfq) if isinstance(rfq, (int, float)) else 0,
        }
    except Exception as e:
        print(f"  Warning: could not read Display Equation in {folder_path}: {e}")
        return None

ALL_TENDERS_DIR   = os.path.join(os.path.dirname(__file__), "data", "All_Tenders")
ETENDERS_DIR      = os.path.join(os.path.dirname(__file__), "data", "etenders.gov.za")
POWER_BI_FILE     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "eTender_PowerBI_Data.xlsx")
PBI_COLUMNS       = ["BATCH_LABEL", "BATCH_ORDER", "BATCH_TYPE", "BATCH_START_DATE", "BATCH_SORT_KEY", "REPORT_DATE", "SOURCE", "SOURCE_SORT", "NNT", "ICT", "RFQ"]
HDR_FONT      = Font(bold=True, color="FFFFFF")
HDR_FILL      = PatternFill("solid", fgColor="1F4E79")


def parse_folder(name):
    # Same-month: (M) 2-5 July 2026
    m = re.match(r"\(([MT])\)\s+(\d+)-(\d+)\s+(\w+)\s+(\d{4})", name)
    if m:
        btype, d1, d2, month_name, year = m.groups()
        end   = datetime.strptime(f"{d2} {month_name} {year}", "%d %B %Y")
        start = end.replace(day=int(d1))
        return btype, start, end
    # Cross-month: (T) 29 Jun-01 Jul 2026
    m = re.match(r"\(([MT])\)\s+(\d+)\s+(\w+)-(\d+)\s+(\w+)\s+(\d{4})", name)
    if m:
        btype, d1, m1, d2, m2, year = m.groups()
        start = datetime.strptime(f"{d1} {m1} {year}", "%d %b %Y")
        end   = datetime.strptime(f"{d2} {m2} {year}", "%d %b %Y")
        return btype, start, end
    return None


all_rows = []

# Build combined folder map: prefer All_Tenders when both dirs have the same batch name
folder_map = {}  # folder_name -> folder_path
for folder_name in sorted(os.listdir(ETENDERS_DIR)):
    fp = os.path.join(ETENDERS_DIR, folder_name)
    if os.path.isdir(fp) and parse_folder(folder_name):
        folder_map[folder_name] = fp
for folder_name in sorted(os.listdir(ALL_TENDERS_DIR)):
    fp = os.path.join(ALL_TENDERS_DIR, folder_name)
    if os.path.isdir(fp) and parse_folder(folder_name):
        folder_map[folder_name] = fp  # overwrites etenders.gov.za if duplicate

for folder_name in sorted(folder_map.keys()):
    folder_path = folder_map[folder_name]
    parsed = parse_folder(folder_name)
    if not parsed:
        print(f"Skipping: {folder_name}")
        continue

    btype, start, end = parsed
    if start.month == end.month:
        batch_label = f"({btype}) {start.day}-{end.day} {end.strftime('%b %y')}"
    else:
        batch_label = f"({btype}) {start.strftime('%d %b')}-{end.strftime('%d %b %y')}"

    # Read eTenders (All) totals from Display Equation — authoritative source
    etenders_all_total = read_etenders_all_from_equation(folder_path)
    if etenders_all_total:
        print(f"  {batch_label}: eTenders (All) from Display Equation — "
              f"NNT={etenders_all_total['NNT']} ICT={etenders_all_total['ICT']} RFQ={etenders_all_total['RFQ']}")
    else:
        print(f"  {batch_label}: no Display Equation found — eTenders (All) will be 0")
        etenders_all_total = {"NNT": 0, "ICT": 0, "RFQ": 0}

    batches_dir = os.path.join(folder_path, "batches")
    current = start
    while current <= end:
        day_str  = current.strftime("%Y_%m_%d")
        day_file = os.path.join(batches_dir, f"tenders_{day_str}.xlsx")
        if os.path.exists(day_file):
            day_df     = pd.read_excel(day_file)
            day_counts = calculate_counts(day_df)
        else:
            day_counts = {src: {"NNT": 0, "ICT": 0, "RFQ": 0} for src, _ in SOURCES}

        # eTenders (All) total recorded only on the last day of the batch
        day_counts["eTenders (All)"] = etenders_all_total if current == end else {"NNT": 0, "ICT": 0, "RFQ": 0}

        for source, _ in SOURCES:
            data = day_counts.get(source, {"NNT": 0, "ICT": 0, "RFQ": 0})
            all_rows.append({
                "BATCH_LABEL":      batch_label,
                "BATCH_TYPE":       btype,
                "BATCH_START_DATE": start,
                "BATCH_SORT_KEY":   -start.toordinal(),
                "REPORT_DATE":      current,
                "SOURCE":           source,
                "SOURCE_SORT":      _SOURCE_SORT.get(source, 999),
                "NNT":              data["NNT"],
                "ICT":              data["ICT"],
                "RFQ":              data["RFQ"],
            })
        current += timedelta(days=1)

    print(f"  {batch_label}: {(end - start).days + 1} days")

combined = pd.DataFrame(all_rows, columns=[c for c in PBI_COLUMNS if c != "BATCH_ORDER"])
combined["REPORT_DATE"] = pd.to_datetime(combined["REPORT_DATE"])
combined.sort_values("REPORT_DATE", ascending=False, inplace=True)
combined.reset_index(drop=True, inplace=True)

# Assign BATCH_ORDER: 1 = newest batch, incrementing for older batches
# Rank by BATCH_START_DATE descending (newest = rank 1)
unique_batches = (
    combined[["BATCH_LABEL", "BATCH_START_DATE"]]
    .drop_duplicates("BATCH_LABEL")
    .sort_values("BATCH_START_DATE", ascending=False)
    .reset_index(drop=True)
)
unique_batches["BATCH_ORDER"] = unique_batches.index + 1
order_map = dict(zip(unique_batches["BATCH_LABEL"], unique_batches["BATCH_ORDER"]))
combined["BATCH_ORDER"] = combined["BATCH_LABEL"].map(order_map)
combined = combined[PBI_COLUMNS]

wb = Workbook()
ws = wb.active
ws.title = "Tender Data"
for col_idx, col_name in enumerate(PBI_COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
for row_idx, row in enumerate(combined.itertuples(index=False), 2):
    for col_idx, col_name in enumerate(PBI_COLUMNS, 1):
        value = getattr(row, col_name)
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if col_name in ("REPORT_DATE", "BATCH_START_DATE") and hasattr(value, "strftime"):
            cell.number_format = "YYYY/MM/DD"

last_col = get_column_letter(len(PBI_COLUMNS))
last_row = len(combined) + 1
tbl = Table(displayName="TenderData", ref=f"A1:{last_col}{last_row}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)
wb.save(POWER_BI_FILE)
print(f"\nDone — {len(combined)} rows written to {POWER_BI_FILE}")
