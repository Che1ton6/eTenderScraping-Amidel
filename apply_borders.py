"""One-off script: applies thin borders to the current equation file."""
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

EQUATION_FILE = (
    r"C:\Users\CheltonGraham\OneDrive - Amidel (Pty) Ltd"
    r"\Documents\Sales\Sales Auto Hub\Scraping and Reports"
    r"\ICT & RFQ\Bhekis conditional Product\RFQ_and_ICT_Equation.xlsx"
)

wb = load_workbook(EQUATION_FILE)
ws = wb["Sheet1"]

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

num_batches = 0
col = 2
while ws.cell(row=1, column=col).value is not None:
    num_batches += 1
    col += 3

total_rows = 28
total_cols = 1 + num_batches * 3

for r in range(1, total_rows + 1):
    for c in range(1, total_cols + 1):
        ws.cell(row=r, column=c).border = border

wb.save(EQUATION_FILE)
print(f"Done — borders applied to {total_rows} rows x {total_cols} cols ({num_batches} batches).")
