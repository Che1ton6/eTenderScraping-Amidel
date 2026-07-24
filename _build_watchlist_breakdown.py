#!/usr/bin/env python3
"""One-off: build enhanced eTenders x Watchlist breakdown for (M) 9-12 July 2026."""
import io, sys, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from BatchProcessor import SOURCES, ICT_CATEGORIES, RFQ_TYPE, _xl_match, _normalize_tender_id

BATCH = os.path.join(HERE, 'data', 'All_Tenders', '(M) 9-12 July 2026')
BATCHES_DIR = os.path.join(BATCH, 'batches')
END_PRODUCT = os.path.join(BATCH, 'end product', 'RFQ_and_ICT_Checker_(9 - 12 July).xlsx')

etenders_frames = []
for f in sorted(os.listdir(BATCHES_DIR)):
    if f.startswith('tenders_') and f.endswith('.xlsx'):
        etenders_frames.append(pd.read_excel(os.path.join(BATCHES_DIR, f)))
etenders_df = pd.concat(etenders_frames, ignore_index=True)
own_df = pd.read_excel(END_PRODUCT, sheet_name='Tender Data')
print(f'eTenders rows: {len(etenders_df)}  |  Own-website rows: {len(own_df)}')


def _match_department(df, pattern):
    if pattern is None:
        return pd.Series([True] * len(df), index=df.index)
    dept = df['DEPARTMENT'].fillna('').astype(str)
    return dept.apply(lambda v: _xl_match(v, pattern))


def _norm_ids(sub_df):
    ids = sub_df['TENDER_ID'].fillna('').astype(str).str.strip()
    return set(_normalize_tender_id(t) for t in ids if t)


def _flags(sub_df):
    cat = sub_df['CATEGORY'].fillna('').astype(str)
    typ = sub_df['TENDER_TYPE'].fillna('').astype(str)
    ict = cat.str.lower().isin([c.lower() for c in ICT_CATEGORIES])
    rfq = typ.str.lower() == RFQ_TYPE.lower()
    return ict, rfq


rows = []
for label, pattern in SOURCES:
    e_sub = etenders_df[_match_department(etenders_df, pattern)]
    w_sub = own_df[_match_department(own_df, pattern)]

    e_ids = _norm_ids(e_sub)
    w_ids = _norm_ids(w_sub)

    e_blank = int((e_sub['TENDER_ID'].fillna('').astype(str).str.strip() == '').sum())
    w_blank = int((w_sub['TENDER_ID'].fillna('').astype(str).str.strip() == '').sum())

    unique_combined = len(e_ids | w_ids) + e_blank + w_blank
    only_in_own = len(w_ids - e_ids) + w_blank
    overlap = len(e_ids & w_ids)

    combined = pd.concat([e_sub, w_sub], ignore_index=True)
    if len(combined):
        tid = combined['TENDER_ID'].fillna('').astype(str).str.strip()
        nid = [_normalize_tender_id(t) if t else f'__blank_{i}' for i, t in enumerate(tid)]
        combined = combined.assign(__nid=nid)
        combined_unique = combined.drop_duplicates(subset=['__nid'])
    else:
        combined_unique = combined
    ict_u, rfq_u = _flags(combined_unique)

    rows.append({
        'Watchlist Entity': label,
        'Pattern': pattern if pattern else '(all rows)',
        'From eTenders': len(e_sub),
        'From Own Website': len(w_sub),
        'Overlap (in both)': overlap,
        'Only in Own (missed by eTenders)': only_in_own,
        'Unique Combined': unique_combined,
        'ICT (unique)': int(ict_u.sum()),
        'RFQs (unique)': int(rfq_u.sum()),
    })

# GRAND TOTAL — dedupe union across every source
e_all_ids = _norm_ids(etenders_df)
w_all_ids = _norm_ids(own_df)
e_blank_all = int((etenders_df['TENDER_ID'].fillna('').astype(str).str.strip() == '').sum())
w_blank_all = int((own_df['TENDER_ID'].fillna('').astype(str).str.strip() == '').sum())
combined_all = pd.concat([etenders_df, own_df], ignore_index=True)
tid_all = combined_all['TENDER_ID'].fillna('').astype(str).str.strip()
nid_all = [_normalize_tender_id(t) if t else f'__blank_{i}' for i, t in enumerate(tid_all)]
combined_all = combined_all.assign(__nid=nid_all).drop_duplicates(subset=['__nid'])
ict_all, rfq_all = _flags(combined_all)

rows.append({
    'Watchlist Entity': 'GRAND TOTAL (all sources, deduped)',
    'Pattern': '-',
    'From eTenders': len(etenders_df),
    'From Own Website': len(own_df),
    'Overlap (in both)': len(e_all_ids & w_all_ids),
    'Only in Own (missed by eTenders)': len(w_all_ids - e_all_ids) + w_blank_all,
    'Unique Combined': len(e_all_ids | w_all_ids) + e_blank_all + w_blank_all,
    'ICT (unique)': int(ict_all.sum()),
    'RFQs (unique)': int(rfq_all.sum()),
})

summary_df = pd.DataFrame(rows)
print()
print(summary_df.to_string(index=False))

# Detail sheet
detail_rows = []
for label, pattern in SOURCES:
    if pattern is None:
        continue
    e_sub = etenders_df[_match_department(etenders_df, pattern)]
    w_sub = own_df[_match_department(own_df, pattern)]
    e_ids = _norm_ids(e_sub)
    w_ids = _norm_ids(w_sub)

    for _, r in e_sub.iterrows():
        tid = str(r.get('TENDER_ID', '') or '').strip()
        nid = _normalize_tender_id(tid) if tid else ''
        origin = 'Both' if nid and nid in w_ids else 'eTenders only'
        detail_rows.append({
            'Watchlist Entity': label, 'Origin': origin,
            'TENDER_ID': r.get('TENDER_ID', ''),
            'PUBLICATION_DATE': r.get('PUBLICATION_DATE', ''),
            'CLOSING_DATE': r.get('CLOSING_DATE', ''),
            'DEPARTMENT': r.get('DEPARTMENT', ''),
            'TENDER_SOURCE': r.get('TENDER_SOURCE', ''),
            'CATEGORY': r.get('CATEGORY', ''),
            'TENDER_TYPE': r.get('TENDER_TYPE', ''),
            'TENDER_DESCRIPTION': r.get('TENDER_DESCRIPTION', ''),
            'LINK': r.get('LINK', ''),
        })
    for _, r in w_sub.iterrows():
        tid = str(r.get('TENDER_ID', '') or '').strip()
        nid = _normalize_tender_id(tid) if tid else ''
        if nid and nid in e_ids:
            continue
        detail_rows.append({
            'Watchlist Entity': label, 'Origin': 'Own website only',
            'TENDER_ID': r.get('TENDER_ID', ''),
            'PUBLICATION_DATE': r.get('PUBLICATION_DATE', ''),
            'CLOSING_DATE': r.get('CLOSING_DATE', ''),
            'DEPARTMENT': r.get('DEPARTMENT', ''),
            'TENDER_SOURCE': r.get('TENDER_SOURCE', ''),
            'CATEGORY': r.get('CATEGORY', ''),
            'TENDER_TYPE': r.get('TENDER_TYPE', ''),
            'TENDER_DESCRIPTION': r.get('TENDER_DESCRIPTION', ''),
            'LINK': r.get('LINK', ''),
        })
detail_df = pd.DataFrame(detail_rows)
print(f'\nDetail rows: {len(detail_df)}')

# Write workbook
out_dir = os.path.join(BATCH, 'eTenders Watchlist Breakdown')
os.makedirs(out_dir, exist_ok=True)
out_file = os.path.join(out_dir, 'eTenders_Watchlist_Breakdown.xlsx')

wb = Workbook()
ws = wb.active
ws.title = 'Summary'

HDR_FONT = Font(bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1C3880')
TOTAL_FILL = PatternFill('solid', fgColor='F5A000')
_thin = Side(style='thin', color='C0C0C0')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

ws['A1'] = 'Batch'
ws['B1'] = os.path.basename(BATCH)
ws['A1'].font = Font(bold=True)
ws['A2'] = 'eTenders (portal) rows'
ws['B2'] = len(etenders_df)
ws['A2'].font = Font(bold=True)
ws['A3'] = 'Own-website (watchlist) rows'
ws['B3'] = len(own_df)
ws['A3'].font = Font(bold=True)
ws['A4'] = 'Unique combined (all sources)'
ws['B4'] = int(summary_df.iloc[-1]['Unique Combined'])
ws['A4'].font = Font(bold=True)

cols = list(summary_df.columns)
HDR_ROW = 6
for c, name in enumerate(cols, 1):
    cell = ws.cell(row=HDR_ROW, column=c, value=name)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = BORDER

for r_idx, row in enumerate(summary_df.itertuples(index=False), HDR_ROW + 1):
    is_total = (row[0] == 'GRAND TOTAL (all sources, deduped)')
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c, value=val)
        cell.border = BORDER
        if c >= 3:
            cell.alignment = Alignment(horizontal='center')
        if is_total:
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL

for col in ws.columns:
    letter = col[0].column_letter
    max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
    ws.column_dimensions[letter].width = min(max_len + 2, 42)
ws.row_dimensions[HDR_ROW].height = 32
ws.freeze_panes = f'A{HDR_ROW + 1}'

ws2 = wb.create_sheet('Matched Tenders (Detail)')
if len(detail_df):
    for c, name in enumerate(detail_df.columns, 1):
        cell = ws2.cell(row=1, column=c, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    for r_idx, row in enumerate(detail_df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c, value=val)
            cell.border = BORDER
    for col in ws2.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws2.column_dimensions[letter].width = min(max_len + 2, 60)
    ws2.freeze_panes = 'A2'
else:
    ws2['A1'] = 'No matches'

wb.save(out_file)
print('\nWROTE:', out_file)
