#!/usr/bin/env python3
"""One-off: patch the (T) 13-15 July 'all_but_etenders' batch so that:
  1. eTenders (All) row shows 0 (no eTenders portal scrape occurred).
  2. A TOTAL row is written below the last source.

Applies to both the Display Equation copy in the batch folder AND the end
product Final tab.
"""
import io, sys, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from BatchProcessor import SOURCES

BATCH = os.path.join(HERE, 'data', 'all_but_etenders', '(T) 13-15 July 2026')
EQUATION = os.path.join(BATCH, 'Display Equation', 'RFQ_and_ICT_Equation.xlsx')
END_PRODUCT = os.path.join(BATCH, 'end product', 'RFQ_and_ICT_Checker_(13 - 15 July).xlsx')

def _backup(p):
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = f"{os.path.splitext(p)[0]}_pre_TOTAL_backup_{ts}.xlsx"
    shutil.copy2(p, dst)
    print(f"  backup: {os.path.basename(dst)}")


def patch_equation(path):
    wb = load_workbook(path)
    ws = wb['Sheet1']
    # (T) 13-15 Jul is the newest batch => columns B/C/D
    # Zero the eTenders (All) row at row 3
    for c in (2, 3, 4):
        ws.cell(row=3, column=c, value=0)
    print(f"  row 3 (eTenders (All)) -> 0 in columns B/C/D")

    total_row = 3 + len(SOURCES)  # row 31 for 28 sources
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for c in (2, 3, 4):
        letter = ws.cell(row=total_row, column=c).column_letter
        ws.cell(row=total_row, column=c,
                value=f"=SUM({letter}4:{letter}{total_row - 1})").font = Font(bold=True)
    print(f"  row {total_row} (TOTAL) -> =SUM formulas in columns B/C/D")

    wb.save(path)


def patch_end_product(path):
    wb = load_workbook(path)
    ws = wb['Final']
    for c in (2, 3, 4):
        ws.cell(row=3, column=c, value=0)
    print(f"  row 3 (eTenders (All)) -> 0")

    total_row = 3 + len(SOURCES)
    # If the sheet already had a summary row here, overwrite with a labelled TOTAL
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for c in (2, 3, 4):
        letter = ws.cell(row=total_row, column=c).column_letter
        ws.cell(row=total_row, column=c,
                value=f"=SUM({letter}4:{letter}{total_row - 1})").font = Font(bold=True)
    print(f"  row {total_row} (TOTAL) -> =SUM formulas")

    wb.save(path)


def main():
    if not os.path.exists(EQUATION) or not os.path.exists(END_PRODUCT):
        print("MISSING file(s) — check paths"); return
    print("Backing up:")
    _backup(EQUATION); _backup(END_PRODUCT)
    print("\nPatching Display Equation:")
    patch_equation(EQUATION)
    print("\nPatching End Product Final tab:")
    patch_end_product(END_PRODUCT)
    print("\nDone.")

if __name__ == '__main__':
    main()
