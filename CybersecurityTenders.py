#!/usr/bin/env python3
"""
Cybersecurity Tenders Generator
Produces 'Cybersecurity Tenders.xlsx' in the batch folder root.
Captures every tender that is (RFQ or ICT-category) AND has a
cybersecurity angle in its description or category.
"""

import os
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from BatchProcessor import BATCH_COLUMNS, ICT_CATEGORIES, RFQ_TYPE, write_tender_rows

# ── Filter keyword lists ───────────────────────────────────────────────────────

_CYBER_KEYWORDS = [
    "cybersecurity", "cyber security", "cyber-security",
    "information security", "infosec", "information assurance",
    "penetration test", "pen test", "pentest", "penetration testing",
    "vulnerability assessment", "vulnerability scan", "vulnerability management",
    "security audit", "security assessment",
    "network security", "endpoint security", "cloud security",
    "firewall", "intrusion detection", "intrusion prevention",
    "siem", "security information and event",
    "data protection", "data security", "data breach", "data loss prevention",
    "security operations centre", "security operations center",
    "incident response", "threat intelligence", "threat detection",
    "ransomware", "malware", "phishing",
    "ethical hacking", "red team", "blue team",
    "iso 27001", "iso27001", "nist cybersecurity",
    "security governance", "cyber governance", "cyber risk",
    "encryption", "cryptography", "public key infrastructure",
    "identity and access management", "privileged access management",
    "zero trust", "devsecops",
    "security awareness training", "cyber awareness",
    "cyber crime", "cybercrime",
]

_ICT_CATEGORIES_LOWER = {c.lower() for c in ICT_CATEGORIES}

# ── Styling ────────────────────────────────────────────────────────────────────

_TITLE_FILL = PatternFill("solid", fgColor="1A3A5C")
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
_HDR_FILL   = PatternFill("solid", fgColor="2E75B6")
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL   = PatternFill("solid", fgColor="DEEAF1")
_THIN       = Side(style="thin", color="BFBFBF")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP       = Alignment(wrap_text=True, vertical="top")


def _is_cyber(row: pd.Series) -> bool:
    desc = str(row.get("TENDER_DESCRIPTION") or "").lower()
    cat  = str(row.get("CATEGORY") or "").lower()
    return any(kw in desc or kw in cat for kw in _CYBER_KEYWORDS)


def _is_rfq(row: pd.Series) -> bool:
    return str(row.get("TENDER_TYPE") or "").lower() == RFQ_TYPE.lower()


def _is_ict(row: pd.Series) -> bool:
    return str(row.get("CATEGORY") or "").lower() in _ICT_CATEGORIES_LOWER


# ── Public entry point ─────────────────────────────────────────────────────────

def create_cybersecurity_tenders(df: pd.DataFrame, batch_folder: str) -> int:
    """
    Filter df for cybersecurity-relevant tenders and write
    Cybersecurity Tenders.xlsx into batch_folder.
    Returns the number of tenders written (0 if none found).
    """
    if df is None or df.empty:
        logging.info("No tenders — skipping Cybersecurity Tenders")
        return 0

    mask = df.apply(
        lambda row: _is_cyber(row) and (_is_rfq(row) or _is_ict(row)), axis=1
    )
    cyber_df = df[mask].copy()

    if cyber_df.empty:
        logging.info("No cybersecurity tenders found — skipping Cybersecurity Tenders")
        return 0

    # Sort: RFQs first, then by closing date ascending
    cyber_df["_is_rfq"] = cyber_df.apply(_is_rfq, axis=1)
    cyber_df.sort_values(["_is_rfq", "CLOSING_DATE"], ascending=[False, True], inplace=True)
    cyber_df.drop(columns=["_is_rfq"], inplace=True)
    cyber_df.reset_index(drop=True, inplace=True)

    filepath = os.path.join(batch_folder, "Cybersecurity Tenders.xlsx")
    n = len(cyber_df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cybersecurity Tenders"

    num_cols = len(BATCH_COLUMNS)

    # ── Row 1: title banner ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title = ws.cell(row=1, column=1,
                    value=f"Amidel (Pty) Ltd — Cybersecurity Tenders   |   {n} tender(s) found")
    title.fill      = _TITLE_FILL
    title.font      = _TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── Row 2: column headers ─────────────────────────────────────────────────
    for col_idx, col_name in enumerate(BATCH_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.alignment = _WRAP
        cell.border    = _BORDER
    ws.row_dimensions[2].height = 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    ordered = cyber_df.reindex(columns=BATCH_COLUMNS)
    for row_idx, row in enumerate(ordered.itertuples(index=False), 3):
        fill = _ALT_FILL if row_idx % 2 == 1 else None
        for col_idx, (col_name, value) in enumerate(zip(BATCH_COLUMNS, row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _WRAP
            cell.border    = _BORDER
            if fill:
                cell.fill = fill
            if col_name == "LINK" and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style     = "Hyperlink"
                cell.alignment = _WRAP
        ws.row_dimensions[row_idx].height = 40

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "REPORT_DATE": 13, "RECORD_ID": 10, "TENDER_ID": 22,
        "PUBLICATION_DATE": 14, "CLOSING_DATE": 14, "CLOSING_TIME": 12,
        "TENDER_TYPE": 22, "TENDER_DESCRIPTION": 55, "TENDER_SOURCE": 20,
        "DEPARTMENT": 30, "PROVINCE": 14, "ESUBMISSION": 12,
        "CATEGORY": 28, "IS_THERE_A_BRIEFING_SESSION": 14,
        "COMPULSORY_BRIEFING": 14, "BRIEFING_SESSION_VENUE": 30,
        "LINK": 30,
    }
    for col_idx, col_name in enumerate(BATCH_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.freeze_panes = "A3"

    wb.save(filepath)
    logging.info(f"Cybersecurity Tenders saved: {filepath} ({n} tender(s))")
    return n
