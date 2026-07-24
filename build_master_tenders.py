"""
One-time seed script: builds data/master_tenders.xlsx from:
  1. DAILYTENDER_REPORT_3.0 (4).xlsx  — TENDER_REPORT (2) sheet (historical data)
  2. All batch files in data/All_Tenders/*/batches/  (new system, June 2026+)
Deduplicates by TENDER_ID. Going forward BatchProcessor appends to this file.
"""

import os
import sys
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from BatchProcessor import TENDER_COLUMNS, HDR_FONT, HDR_FILL, DATE_COLUMNS, _normalize_tender_id

DAILYTENDER_SOURCE = r"C:\Users\CheltonGraham\Downloads\DAILYTENDER_REPORT_3.0 (4).xlsx"
ALL_TENDERS_DIR    = os.path.join(os.path.dirname(__file__), "data", "All_Tenders")
MASTER_FILE        = os.path.join(os.path.dirname(__file__), "data", "master_tenders.xlsx")

# Columns in TENDER_REPORT (2) that map to our standard schema
OLD_COL_MAP = {
    "REPORT_DATE":               "REPORT_DATE",
    "TENDER_ID":                 "TENDER_ID",
    "PUBLICATION_DATE":          "PUBLICATION_DATE",
    "CLOSING_DATE":              "CLOSING_DATE",
    "CLOSING_TIME":              "CLOSING_TIME",
    "TENDER_TYPE":               "TENDER_TYPE",
    "TENDER_DESCRIPTION":        "TENDER_DESCRIPTION",
    "TENDER_SOURCE":             "TENDER_SOURCE",
    "DEPARTMENT ":               "DEPARTMENT",       # trailing space in old file
    "PROVINCE":                  "PROVINCE",
    "ESUBMISSION":               "ESUBMISSION",
    "CATEGORY":                  "CATEGORY",
    "IS_THERE_A_BRIEFING_SESSION": "IS_THERE_A_BRIEFING_SESSION",
    "BRIEFING_DATE":             "BRIEFING_DATE",
    "COMPULSORY_BRIEFING":       "COMPULSORY_BRIEFING",
    "BRIEFING_SESSION_VENUE":    "BRIEFING_SESSION_VENUE",
    "LINK":                      "LINK",
    "SOE":                       "SOE",
    "Cost of Sales Estimate":    "COST_OF_SALES_ESTIMATE",
    "CAPABILITY_AVAILABLE":      "CAPABILITY_AVAILABLE",
    "CAPABILITY_GROUP":          "CAPABILITY_GROUP",
    "REQUIREMENTS":              "REQUIREMENTS",
}

OUTPUT_COLUMNS = [c for c in TENDER_COLUMNS if c != "RECORD_ID"]


def _is_formula(val):
    return isinstance(val, str) and val.strip().startswith("=")


def load_old_data():
    print(f"Loading historical data from {DAILYTENDER_SOURCE} ...")
    df = pd.read_excel(DAILYTENDER_SOURCE, sheet_name="TENDER_REPORT (2)", dtype=str)

    # Drop formula rows (rows where TENDER_ID looks like a formula or is blank)
    df = df[df["TENDER_ID"].notna()]
    df = df[~df["TENDER_ID"].str.startswith("=", na=False)]
    df = df[df["TENDER_ID"].str.strip() != ""]

    # Rename columns to standard schema
    df = df.rename(columns=OLD_COL_MAP)

    # Drop columns not in our schema
    df = df[[c for c in OUTPUT_COLUMNS if c in df.columns]]

    # Fill missing standard columns with None
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df = df[OUTPUT_COLUMNS]
    print(f"  Historical rows loaded: {len(df)}")
    return df


def load_batch_data():
    print(f"Loading new batch data from {ALL_TENDERS_DIR} ...")
    frames = []
    for folder_name in sorted(os.listdir(ALL_TENDERS_DIR)):
        batches_dir = os.path.join(ALL_TENDERS_DIR, folder_name, "batches")
        if not os.path.isdir(batches_dir):
            continue
        for fname in sorted(os.listdir(batches_dir)):
            if not fname.endswith(".xlsx") or "No Tenders" in fname:
                continue
            fpath = os.path.join(batches_dir, fname)
            try:
                df = pd.read_excel(fpath, dtype=str)
                frames.append(df)
            except Exception as e:
                print(f"  Warning: could not read {fpath}: {e}")

    if not frames:
        print("  No batch files found.")
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    combined = pd.concat(frames, ignore_index=True)
    # Normalise column names
    combined.columns = [c.strip() for c in combined.columns]
    if "COST_OF_SALES_ESTIMATE" not in combined.columns and "Cost of Sales Estimate" in combined.columns:
        combined = combined.rename(columns={"Cost of Sales Estimate": "COST_OF_SALES_ESTIMATE"})

    combined = combined[[c for c in OUTPUT_COLUMNS if c in combined.columns]]
    for col in OUTPUT_COLUMNS:
        if col not in combined.columns:
            combined[col] = None
    combined = combined[OUTPUT_COLUMNS]
    print(f"  New batch rows loaded: {len(combined)}")
    return combined


def deduplicate(df):
    seen = set()
    keep = []
    dupes = 0
    for idx, row in df.iterrows():
        tid = str(row.get("TENDER_ID") or "").strip()
        norm = _normalize_tender_id(tid) if tid else ""
        if norm and norm in seen:
            dupes += 1
            continue
        keep.append(idx)
        if norm:
            seen.add(norm)
    print(f"  Deduplication removed {dupes} duplicates — {len(keep)} unique tenders remain.")
    return df.loc[keep].reset_index(drop=True)


def write_master(df):
    os.makedirs(os.path.dirname(MASTER_FILE), exist_ok=True)

    # Assign RECORD_IDs
    df = df.copy()
    df.insert(1, "RECORD_ID", range(1, len(df) + 1))
    all_cols = ["REPORT_DATE", "RECORD_ID"] + [c for c in OUTPUT_COLUMNS if c not in ("REPORT_DATE",)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Tender Data"

    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    for row_idx, row in enumerate(df[all_cols].itertuples(index=False), 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            value = getattr(row, col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in DATE_COLUMNS and value and not pd.isna(value):
                try:
                    if not isinstance(value, datetime):
                        value = pd.to_datetime(value)
                    cell.value = value.to_pydatetime()
                    cell.number_format = "YYYY/MM/DD"
                except Exception:
                    pass

    last_col = get_column_letter(len(all_cols))
    last_row = len(df) + 1
    tbl = Table(displayName="MasterTenders", ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    wb.save(MASTER_FILE)
    print(f"\nMaster file saved: {MASTER_FILE}")
    print(f"Total rows: {len(df)}")


if __name__ == "__main__":
    old_df   = load_old_data()
    new_df   = load_batch_data()

    # Old data first (historical), new data appended — then deduplicate
    # New data takes priority on duplicates (more complete fields)
    combined = pd.concat([new_df, old_df], ignore_index=True)
    combined = deduplicate(combined)

    # Sort by REPORT_DATE descending
    combined["REPORT_DATE"] = pd.to_datetime(combined["REPORT_DATE"], errors="coerce")
    combined.sort_values("REPORT_DATE", ascending=False, inplace=True)
    combined.reset_index(drop=True, inplace=True)

    write_master(combined)
