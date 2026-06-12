#!/usr/bin/env python3
"""
Scraper for EC DPW (Eastern Cape Department of Public Works) tenders.
URL: https://www.ecdpw.gov.za/tenders/
Scrapes all Open tenders with closing date >= 14 days from today.
Output uses the same TENDER_COLUMNS structure as the eTenders scraper.
"""

import logging
import time
import os
import re
from datetime import datetime, timedelta, date
from typing import List, Dict, Any, Optional, Callable

import requests
import fitz          # PyMuPDF
import pytesseract
from PIL import Image
import pandas as pd

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException
)

from BatchProcessor import TENDER_COLUMNS

SOURCE_URL    = "https://www.ecdpw.gov.za/tenders/"
TENDER_SOURCE = "ECDPW.GOV.ZA"
OUTPUT_DIR    = os.path.join("data", "ecdpw")


class ECDPWScraper:

    def __init__(self, log_queue=None, pub_date_from: Optional[date] = None,
                 filter_label: Optional[str] = None):
        self._setupLogging(log_queue)
        self.driver       = None
        self.tenderData: List[Dict[str, Any]] = []
        self.minClosingDate = date.today()
        self.pub_date_from  = pub_date_from
        self.filter_label   = filter_label  # e.g. "2026-05" or "2026" or None
        if pub_date_from:
            logging.info(f"ECDPWScraper initialised — pub date filter: from {pub_date_from}")
        else:
            logging.info("ECDPWScraper initialised — no pub date filter (full scrape)")

    # ── Logging ───────────────────────────────────────────────────────────────

    def _setupLogging(self, log_queue):
        root = logging.getLogger()
        if not root.handlers:
            os.makedirs("logs", exist_ok=True)
            root.setLevel(logging.INFO)
            root.addHandler(logging.FileHandler("logs/scraper.log"))

        if log_queue is not None:
            class _QueueHandler(logging.Handler):
                def __init__(self, q):
                    super().__init__()
                    self.q = q
                def emit(self, record):
                    self.q.put(self.format(record))

            qh = _QueueHandler(log_queue)
            qh.setFormatter(logging.Formatter(
                "%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S"))
            root.addHandler(qh)

    # ── Browser ───────────────────────────────────────────────────────────────

    def setupBrowser(self):
        opts = Options()
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        self.driver = webdriver.Chrome(options=opts)
        logging.info("Browser setup completed")

    def cleanup(self):
        if self.driver:
            try:
                self.driver.quit()
                logging.info("Browser closed successfully")
            except Exception as e:
                logging.warning(f"Error closing browser: {e}")

    # ── Page loading ──────────────────────────────────────────────────────────

    def waitForTableData(self, timeout: int = 45):
        """
        Block until the table has at least one real tender row.
        A real row has 4+ cells and its first cell does NOT contain newlines
        (which would indicate it's the filter/dropdown row).
        """
        try:
            WebDriverWait(self.driver, timeout).until(
                lambda d: any(
                    len(r.find_elements(By.TAG_NAME, "td")) >= 4
                    and "\n" not in (r.find_elements(By.TAG_NAME, "td")[0].text or "")
                    and (r.find_elements(By.TAG_NAME, "td")[0].text or "").strip() not in ("", "Loading...")
                    for r in d.find_elements(By.XPATH, "//table/tbody/tr")
                    if r.is_displayed()
                )
            )
        except TimeoutException:
            logging.warning(f"Table did not populate within {timeout}s — proceeding anyway")

    def navigateToPage(self):
        logging.info(f"Navigating to {SOURCE_URL}")
        self.driver.get(SOURCE_URL)
        time.sleep(5)
        try:
            self.driver.execute_script("""
                let modals = document.querySelectorAll('.modal.show');
                modals.forEach(m => m.parentNode.removeChild(m));
                let backdrop = document.querySelector('.modal-backdrop');
                if (backdrop) backdrop.parentNode.removeChild(backdrop);
                document.body.classList.remove('modal-open');
            """)
        except Exception:
            pass
        # Dismiss cookie consent banner if present
        try:
            self.driver.execute_script("""
                let el = document.querySelector('.cky-notice-group, .cky-consent-bar, #cookie-law-info-bar');
                if (el) el.remove();
                let overlay = document.querySelector('.cky-overlay');
                if (overlay) overlay.remove();
            """)
        except Exception:
            pass
        self.waitForTableData()
        logging.info("Page loaded successfully")

    # ── Filters ───────────────────────────────────────────────────────────────

    def setFilters(self):
        """Set Status = Open."""
        for sel_el in self.driver.find_elements(By.TAG_NAME, "select"):
            options = [o.text.strip() for o in sel_el.find_elements(By.TAG_NAME, "option")]
            if "Open" in options:
                if Select(sel_el).first_selected_option.text.strip() != "Open":
                    Select(sel_el).select_by_visible_text("Open")
                    time.sleep(2)
                    self.waitForTableData()
                logging.info("Status filter set to Open")
                return
        logging.warning("Status filter not found — scraping all visible rows")

    # ── Row parsing ───────────────────────────────────────────────────────────

    def _parseDate(self, text: str) -> str:
        """Return YYYY/MM/DD or empty string if no format matches."""
        text = (text or "").strip()
        if not text:
            return ""
        # Try full string first (handles "18 JUNE 2026"), then first token only
        # (handles "2026-06-09 13:27:49" by stripping the time suffix).
        first_token = text.split()[0]
        for candidate in [text, text.title(), first_token]:
            for fmt in ("%d/%m/%Y", "%d %B %Y", "%Y-%m-%d", "%d %b %Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(candidate, fmt).strftime("%Y/%m/%d")
                except ValueError:
                    continue
        return ""

    def _parseTime(self, text: str) -> str:
        """Normalise time strings like '11H00 am', '11:00', '11H00' → 'HH:MM'."""
        text = (text or "").strip()
        # Replace H separator: "11H00" → "11:00"
        text = re.sub(r'(\d{1,2})[Hh](\d{2})', r'\1:\2', text)
        # Strip am/pm suffix for storage
        text = re.sub(r'\s*(am|pm)$', '', text, flags=re.IGNORECASE).strip()
        # Validate HH:MM
        if re.match(r'^\d{1,2}:\d{2}$', text):
            return text
        return ""

    def _parseClosingDateTime(self, text: str):
        """Split 'DD/MM/YYYY HH:MM' into (date_str, time_str)."""
        parts = (text or "").strip().split()
        date_str = self._parseDate(parts[0]) if parts else ""
        time_str = parts[1] if len(parts) > 1 else ""
        return date_str, time_str

    _TYPE_MAP = {
        "REQUEST FOR QUOTATION":    "Request for Quotation",
        "QUOTATION REQUEST":        "Request for Quotation",
        "PRICE QUOTATION":          "Request for Quotation",
        "REQUEST FOR BID":          "Request for Bid",
        "BID REQUEST":              "Request for Bid",
        "INVITATION TO BID":        "Request for Bid",
        "REQUEST FOR PROPOSAL":     "Request for Proposal",
        "PROPOSAL REQUEST":         "Request for Proposal",
        "REQUEST FOR INFORMATION":  "Request for Information",
        "EXPRESSION OF INTEREST":   "Expression of Interest",
    }

    # Keywords that reliably indicate RFP (professional consulting services)
    _RFP_KEYWORDS = [
        "PROFESSIONAL CIVIL", "PROFESSIONAL STRUCTURAL", "CIVIL AND STRUCTURAL",
        "PROFESSIONAL ARCHITECTURAL", "ARCHITECTURAL SERVICES", "ARCHITECTURAL AND PRINCIPAL",
        "PROFESSIONAL QUANTITY SURVEYING", "QUANTITY SURVEYING SERVICES",
        "PROFESSIONAL ENGINEERING", "ENGINEERING SERVICES",
        "PROFESSIONAL MECHANICAL", "MECHANICAL ENGINEERING SERVICES",
        "PROFESSIONAL ELECTRICAL", "ELECTRICAL ENGINEERING SERVICES",
        "PROFESSIONAL GEOTECHNICAL", "PROFESSIONAL TOWN PLANNING",
        "PROFESSIONAL LANDSCAPE", "PROFESSIONAL PROJECT MANAGEMENT",
        "PRINCIPAL AGENT", "PRINCIPAL AGENCY",
        "HEALTH AND SAFETY AGENT", "OCCUPATIONAL HEALTH AND SAFETY AGENT",
        "CONSTRUCTION HEALTH AND SAFETY AGENT",
    ]

    # Keywords that reliably indicate RFQ (goods / small-value services)
    _RFQ_KEYWORDS = [
        "QUOTATION", "PRICE QUOTATION",
        "SUPPLY AND DELIVER", "SUPPLY & DELIVERY", "SUPPLY, DELIVER",
        "SUPPLY AND DELIVERY", "PURCHASING OF",
        "CATERING", "ACCOMMODATION AND MEALS",
        "HIRING OF OCCASIONAL",
    ]

    def _inferTypeFromDescription(self, description: str) -> str:
        """
        Infer tender type from the tender description when OCR cannot determine it.
        Returns one of the standard type strings, or empty string if unsure.
        """
        desc = (description or "").upper()

        if "EXPRESSION OF INTEREST" in desc:
            return "Expression of Interest"

        if any(kw in desc for kw in self._RFQ_KEYWORDS):
            return "Request for Quotation"

        if any(kw in desc for kw in self._RFP_KEYWORDS):
            return "Request for Proposal"

        # Construction, maintenance, security, term contracts → RFB
        return "Request for Bid"

    def _extractFromPdf(self, url: str) -> Dict[str, str]:
        """
        Download a tender PDF, OCR pages 1-2, and extract closing date,
        closing time, and tender type. Returns a dict with any fields found.
        """
        result = {}
        try:
            url = url.replace("http://", "https://")
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            doc = fitz.open(stream=resp.content, filetype="pdf")
            mat = fitz.Matrix(200 / 72, 200 / 72)  # ~200 DPI

            full_text = ""
            for page_idx in range(min(2, len(doc))):
                pix = doc[page_idx].get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                page_text = pytesseract.image_to_string(img)
                # If page looks upside-down (no common keywords), try 180° rotation
                if "CLOSING" not in page_text.upper() and "DATE" not in page_text.upper():
                    page_text = pytesseract.image_to_string(img.rotate(180))
                full_text += page_text + "\n"

            upper_text = full_text.upper()

            # ── Closing date ──────────────────────────────────────────────────
            # Handles variants:
            #   "CLOSING DATE: 18 JUNE 2026 CLOSING TIME: 11H00 am"
            #   "CLOSING DATE: | 07 July 2026 CLOSING TIME: 11H00"
            #   "Closing date and time: 07 July 2026 at 11:00 am"
            date_keyword_pos = upper_text.find("CLOSING DATE")
            if date_keyword_pos != -1 and "CLOSING_DATE" not in result:
                snippet = upper_text[date_keyword_pos: date_keyword_pos + 120]
                # Find DD MONTH YYYY pattern
                dm = re.search(r'(\d{1,2})\s+([A-Z]{3,9})\s+(\d{4})', snippet)
                if dm:
                    parsed = self._parseDate(f"{dm.group(1)} {dm.group(2).title()} {dm.group(3)}")
                    if parsed:
                        result["CLOSING_DATE"] = parsed
                else:
                    # Try DD/MM/YYYY
                    dm2 = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', snippet)
                    if dm2:
                        parsed = self._parseDate(dm2.group(1))
                        if parsed:
                            result["CLOSING_DATE"] = parsed

            # ── Closing time ──────────────────────────────────────────────────
            time_keyword_pos = upper_text.find("CLOSING TIME")
            if time_keyword_pos == -1:
                time_keyword_pos = upper_text.find("CLOSING DATE AND TIME")
            if time_keyword_pos != -1 and "CLOSING_TIME" not in result:
                snippet = upper_text[time_keyword_pos: time_keyword_pos + 80]
                tm = re.search(r'(\d{1,2}[H:]\d{2})', snippet)
                if tm:
                    t = self._parseTime(tm.group(1))
                    if t:
                        result["CLOSING_TIME"] = t

            # ── Tender type ───────────────────────────────────────────────────
            if "TENDER_TYPE" not in result:
                for phrase, mapped in self._TYPE_MAP.items():
                    if phrase in upper_text:
                        result["TENDER_TYPE"] = mapped
                        break

        except Exception as e:
            logging.warning(f"PDF extract failed for {url}: {e}")
        return result

    def processTenderRow(self, cells) -> Optional[Dict[str, Any]]:
        """
        Parse one table row into a tender dict using TENDER_COLUMNS structure.
        Columns: Status | Tender Details | Tender Dates | Tender Documents
        """
        if len(cells) < 3:
            return None

        # Skip the filter row — its Status cell contains newlines (dropdown options)
        if "\n" in (cells[0].text or ""):
            return None

        tender = {col: "" for col in TENDER_COLUMNS if col != "RECORD_ID"}
        tender["REPORT_DATE"]   = date.today().strftime("%Y/%m/%d")
        tender["TENDER_SOURCE"] = TENDER_SOURCE
        tender["PROVINCE"]      = "Eastern Cape"
        tender["DEPARTMENT"]    = "Eastern Cape Department of Public Works and Infrastructure"

        # Col 0 is "Status" (e.g. "Open for Bidding") — not a tender type.
        # EC DPW listing does not expose RFQ/RFB categorisation.

        # ── Col 1: Tender Details ─────────────────────────────────────────────
        # Structure: DESCRIPTION\n\nRegion: X\nBid Number: Y
        lines = [l.strip() for l in cells[1].text.splitlines() if l.strip()]
        desc_lines = []
        for line in lines:
            lower = line.lower()
            if lower.startswith("bid number:") or lower.startswith("bid no:"):
                tender["TENDER_ID"] = line.split(":", 1)[-1].strip()
            elif lower.startswith("region:"):
                tender["REGION"] = line.split(":", 1)[-1].strip()
            elif lower.startswith("department:"):
                tender["DEPARTMENT"] = line.split(":", 1)[-1].strip()
            else:
                desc_lines.append(line)
        tender["TENDER_DESCRIPTION"] = " ".join(desc_lines)
        # Fall back: use description as TENDER_ID if no Bid Number found
        if not tender["TENDER_ID"] and desc_lines:
            tender["TENDER_ID"] = desc_lines[0]

        # ── Col 2: Tender Dates ───────────────────────────────────────────────
        for line in cells[2].text.splitlines():
            line  = line.strip()
            lower = line.lower()
            val   = line.split(":", 1)[-1].strip() if ":" in line else line
            if "clos" in lower:
                tender["CLOSING_DATE"], tender["CLOSING_TIME"] = self._parseClosingDateTime(val)
            elif "advert" in lower or "pub" in lower:
                tender["PUBLICATION_DATE"] = self._parseDate(val)

        # ── Col 3: Tender Documents → LINK ────────────────────────────────────
        if len(cells) > 3:
            links = cells[3].find_elements(By.TAG_NAME, "a")
            if links:
                tender["LINK"] = links[0].get_attribute("href") or ""

        return tender

    def _parsePubDate(self, raw: str) -> Optional[date]:
        raw = str(raw or "").strip()[:10]
        for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        return None

    def applyPublicationDateFilter(self):
        """Remove tenders published before self.pub_date_from (if set)."""
        if not self.pub_date_from:
            return
        before = len(self.tenderData)
        self.tenderData = [
            t for t in self.tenderData
            if (self._parsePubDate(t.get("PUBLICATION_DATE", "")) or date.min) >= self.pub_date_from
        ]
        removed = before - len(self.tenderData)
        if removed:
            logging.info(f"Removed {removed} tenders published before {self.pub_date_from}")

    def _passesClosingFilter(self, tender: Dict[str, Any]) -> bool:
        """Include all tenders during initial scrape; closing dates fetched separately."""
        return True

    def deduplicateByTenderId(self):
        """Keep only the first occurrence of each TENDER_ID."""
        seen, unique = set(), []
        for t in self.tenderData:
            key = t.get("TENDER_ID", "")
            if key not in seen:
                seen.add(key)
                unique.append(t)
        removed = len(self.tenderData) - len(unique)
        if removed:
            logging.info(f"Removed {removed} duplicate tenders")
        self.tenderData = unique

    def fetchClosingDates(self, status_callback: Optional[Callable] = None):
        """
        Navigate to each tender's detail page to extract the closing date.
        Called after scrapeAllPages(), once pagination is complete.
        """
        total = len(self.tenderData)
        logging.info(f"Fetching closing dates from {total} detail pages — DO NOT close the program…")
        for idx, tender in enumerate(self.tenderData, 1):
            link = tender.get("LINK", "")
            if not link:
                continue
            if status_callback:
                status_callback(f"EC DPW — fetching detail page {idx}/{total}…")
            if idx % 10 == 0 or idx == 1:
                logging.info(f"Detail pages progress: {idx}/{total}…")
            try:
                self.driver.get(link)
                time.sleep(2)
                body_text = self.driver.find_element(By.TAG_NAME, "body").text
                for line in body_text.splitlines():
                    lower = line.lower()
                    if "clos" in lower and ("date" in lower or ":" in line):
                        val = line.split(":", 1)[-1].strip() if ":" in line else ""
                        parts = val.split()
                        parsed = self._parseDate(parts[0]) if parts else ""
                        if parsed:
                            tender["CLOSING_DATE"] = parsed
                            tender["CLOSING_TIME"] = parts[1] if len(parts) > 1 else tender["CLOSING_TIME"]
                            logging.info(f"  {tender['TENDER_ID']} | Closing: {parsed}")
                            break
            except Exception as e:
                logging.warning(f"Detail page error for {tender.get('TENDER_ID', '')}: {e}")
        logging.info(f"Detail pages complete: {total}/{total} fetched.")

    def fetchPdfData(self, status_callback: Optional[Callable] = None):
        """Download each tender's PDF and fill in closing date, closing time, tender type."""
        total = len(self.tenderData)
        logging.info(f"Reading {total} PDF tender documents for closing dates — DO NOT close the program…")
        for idx, tender in enumerate(self.tenderData, 1):
            link = tender.get("LINK", "")
            if not link:
                continue
            if status_callback:
                status_callback(f"EC DPW — reading PDF {idx}/{total}…")
            if idx % 10 == 0 or idx == 1:
                logging.info(f"PDF progress: {idx}/{total}…")
            extracted = self._extractFromPdf(link)
            for field, value in extracted.items():
                if not tender.get(field):
                    tender[field] = value
            # Fallback: infer type from description if OCR didn't find it
            if not tender.get("TENDER_TYPE"):
                tender["TENDER_TYPE"] = self._inferTypeFromDescription(
                    tender.get("TENDER_DESCRIPTION", "")
                )
        # Final pass: any tender still missing a type gets inference applied
        for tender in self.tenderData:
            if not tender.get("TENDER_TYPE"):
                tender["TENDER_TYPE"] = self._inferTypeFromDescription(
                    tender.get("TENDER_DESCRIPTION", "")
                )
        logging.info(f"PDF reading complete: {total}/{total} done.")

    def applyClosingDateFilter(self):
        """
        Remove tenders whose closing date is known and < 14 days from today.
        Also remove tenders with no closing date whose publication date is
        more than 6 months old — these are stale records with deleted PDFs.
        """
        before = len(self.tenderData)
        stale_cutoff = date.today() - timedelta(days=180)

        def passes(t):
            # Drop tenders whose closing date has already passed
            raw_close = t.get("CLOSING_DATE", "")
            if raw_close:
                try:
                    return datetime.strptime(raw_close, "%Y/%m/%d").date() >= self.minClosingDate
                except ValueError:
                    pass
            # No closing date — keep only if publication date is recent (drop stale old records)
            raw_pub = t.get("PUBLICATION_DATE", "")
            if raw_pub:
                try:
                    pub_date = datetime.strptime(str(raw_pub)[:10], "%Y-%m-%d").date()
                    return pub_date >= stale_cutoff
                except ValueError:
                    pass
            return True

        self.tenderData = [t for t in self.tenderData if passes(t)]
        removed = before - len(self.tenderData)
        if removed:
            logging.info(f"Removed {removed} tenders (closing date < {self.minClosingDate} or stale with no closing date)")

    # ── Pagination ────────────────────────────────────────────────────────────

    def goToNextPage(self) -> bool:
        """
        Click the next page-number link. Returns False when on the last page.
        The EC DPW site uses <a class="page-number"> links; the current page
        has the additional class "page-current".
        """
        for attempt in range(3):
            try:
                current_els = self.driver.find_elements(By.CSS_SELECTOR, "a.page-number.page-current")
                if not current_els:
                    return False
                current_num = int(current_els[0].text.strip())
                next_num    = current_num + 1
                all_links   = self.driver.find_elements(By.CSS_SELECTOR, "a.page-number")
                next_links  = [el for el in all_links if el.text.strip() == str(next_num)]
                if not next_links:
                    logging.info(f"Page {current_num} is the last page")
                    return False
                self.driver.execute_script("arguments[0].click();", next_links[0])
                time.sleep(2)
                self.waitForTableData()
                return True
            except StaleElementReferenceException:
                time.sleep(0.5)
                continue
            except Exception as e:
                logging.warning(f"Pagination error: {e}")
                return False
        return False

    # ── Scraping ──────────────────────────────────────────────────────────────

    def scrapeAllPages(self, status_callback: Optional[Callable] = None):
        page = 1
        while True:
            logging.info(f"--- Scraping page {page} ---")
            if status_callback:
                status_callback(f"EC DPW — page {page}…")

            rows = self.driver.find_elements(By.XPATH, "//table/tbody/tr")
            real_rows = [
                r for r in rows
                if r.is_displayed()
                and not r.get_attribute("innerHTML").strip().startswith("<td colspan")
            ]

            if not real_rows:
                logging.warning(f"No rows found on page {page}")
                break

            page_tenders = []
            for row in real_rows:
                try:
                    cells  = row.find_elements(By.TAG_NAME, "td")
                    tender = self.processTenderRow(cells)
                    if tender:
                        self.tenderData.append(tender)
                        page_tenders.append(tender)
                        logging.info(f"Tender found: {tender['TENDER_ID']}")
                except StaleElementReferenceException:
                    logging.warning(f"Stale element on page {page} — skipping row")

            # Early stop: if all tenders on this page are older than the pub date filter,
            # tenders on subsequent pages will be even older — no need to continue.
            if self.pub_date_from and page_tenders:
                page_dates = [
                    d for d in (self._parsePubDate(t.get("PUBLICATION_DATE", ""))
                                for t in page_tenders)
                    if d is not None
                ]
                if page_dates and max(page_dates) < self.pub_date_from:
                    logging.info(
                        f"Page {page}: newest pub date ({max(page_dates)}) is before "
                        f"filter date ({self.pub_date_from}) — stopping early"
                    )
                    break

            if not self.goToNextPage():
                break
            page += 1

        logging.info(f"Scraping complete — {len(self.tenderData)} tenders collected")

    # ── Export ────────────────────────────────────────────────────────────────

    def exportToExcel(self) -> str:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        today_str = date.today().strftime("%Y-%m-%d")
        if self.filter_label:
            filename = f"ECDPW_tenders_{self.filter_label}_{today_str}.xlsx"
        else:
            filename = f"ECDPW_tenders_{today_str}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)

        cols_out = [c for c in TENDER_COLUMNS if c != "RECORD_ID"]
        data = [{c: t.get(c, "") for c in cols_out} for t in self.tenderData]
        df = pd.DataFrame(data)[cols_out]
        df.to_excel(filepath, index=False, engine="openpyxl")
        self._saveExcelWithFormatting(filepath, df)
        logging.info(f"EC DPW output saved: {filepath} ({len(df)} tenders)")
        return filepath

    def _saveExcelWithFormatting(self, filepath: str, df: pd.DataFrame):
        """Apply date formatting and hyperlinks — same approach as TenderScraper."""
        try:
            wb = load_workbook(filepath)
            ws = wb.active

            date_style = NamedStyle(name="date_style")
            date_style.number_format = "YYYY/MM/DD"

            date_columns    = {"REPORT_DATE", "PUBLICATION_DATE", "CLOSING_DATE", "BRIEFING_DATE"}
            date_col_indices = [
                df.columns.get_loc(c) + 1
                for c in date_columns if c in df.columns
            ]
            link_col_idx = df.columns.get_loc("LINK") + 1 if "LINK" in df.columns else None

            red_fill   = PatternFill("solid", fgColor="FF0000")
            red_font   = Font(color="FFFFFF", bold=True)
            green_fill = PatternFill("solid", fgColor="C6EFCE")
            green_font = Font(color="276221", bold=True)
            today    = date.today()
            close_col_idx = df.columns.get_loc("CLOSING_DATE") + 1 if "CLOSING_DATE" in df.columns else None
            type_col_idx  = df.columns.get_loc("TENDER_TYPE") + 1 if "TENDER_TYPE" in df.columns else None

            for row_idx in range(2, len(df) + 2):
                for col_idx in date_col_indices:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        try:
                            cell.value = datetime.strptime(cell.value, "%Y/%m/%d")
                            cell.style = date_style
                        except ValueError:
                            pass

                # Colour-code closing date
                if close_col_idx:
                    ccell = ws.cell(row=row_idx, column=close_col_idx)
                    cv = ccell.value
                    if cv:
                        cd = cv.date() if isinstance(cv, datetime) else None
                        if cd:
                            days = (cd - today).days
                            if days < 14:
                                ccell.fill = red_fill
                                ccell.font = red_font

                # Green highlight for Request for Quotation rows
                if type_col_idx:
                    tcell = ws.cell(row=row_idx, column=type_col_idx)
                    if tcell.value == "Request for Quotation":
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = green_fill
                        tcell.font = green_font

                if link_col_idx:
                    cell = ws.cell(row=row_idx, column=link_col_idx)
                    if cell.value and str(cell.value).startswith("http"):
                        cell.hyperlink = cell.value
                        cell.style     = "Hyperlink"

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            wb.save(filepath)
            wb.close()
        except Exception as e:
            logging.warning(f"Excel formatting error: {e}")

    # ── Entry point ───────────────────────────────────────────────────────────

    def run(self, status_callback: Optional[Callable] = None) -> str:
        """Full scrape pipeline. Returns the output file path, or '' if nothing found."""
        try:
            logging.info("Starting EC DPW scraping process")
            self.setupBrowser()
            self.navigateToPage()
            self.setFilters()
            self.scrapeAllPages(status_callback=status_callback)
            self.deduplicateByTenderId()
            self.applyPublicationDateFilter()
            self.fetchPdfData(status_callback=status_callback)
            self.applyClosingDateFilter()
            if self.tenderData:
                filepath = self.exportToExcel()
                logging.info(
                    f"========================================\n"
                    f"  EC DPW SCRAPING 100% COMPLETE\n"
                    f"  {len(self.tenderData)} tenders saved to: {filepath}\n"
                    f"  You may now close the program.\n"
                    f"========================================"
                )
                if status_callback:
                    status_callback(f"EC DPW — 100% complete — {len(self.tenderData)} tenders saved.")
                return filepath
            logging.warning("No tenders found — nothing to save.")
            return ""
        except Exception as e:
            logging.error(f"EC DPW scraping failed: {e}")
            raise
        finally:
            self.cleanup()
