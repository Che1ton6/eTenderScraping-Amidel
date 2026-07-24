#!/usr/bin/env python3
"""One-off: replace ONLY phantom-scraper watchlist counts with 'TBD' for
the (M) 9-12 July 2026 batch.

Scope of change:
  - Restore the batch files from the pre-TBD backups first (undo the earlier
    over-aggressive TBD pass).
  - TBD only the 11 rows where the underlying scraper produced rows without
    valid IDs/publication dates: Matatiele, Ntabankulu, Umzimvubu, Winnie
    Mandela, Mnquma, Great Kei, Amahlathi, GPL, Nelson Mandela Bay, Buffalo
    City, GDoH.
  - Leave eTenders (All), Eskom, Transnet, SITA, RAF, W JHB, SIU, EC DPW
    and the legitimate zeros at their scraped values so the report has
    something to work with.

Scope: batch-folder copies ONLY. Master equation file at
'Sales Auto Hub\...\Bhekis conditional Product\RFQ_and_ICT_Equation.xlsx'
is not touched.
"""
import io, sys, os, shutil, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
BATCH = os.path.join(HERE, 'data', 'All_Tenders', '(M) 9-12 July 2026')
EQUATION = os.path.join(BATCH, 'Display Equation', 'RFQ_and_ICT_Equation.xlsx')
END_PRODUCT = os.path.join(BATCH, 'end product', 'RFQ_and_ICT_Checker_(9 - 12 July).xlsx')

# Rows to TBD (Excel row numbers; both files share the same layout).
# ONLY phantom scrapers where the row-level data is untrustworthy.
PHANTOM_ROWS = {
    8:  'Matatiele LM',
    9:  'Ntabankulu LM',
    10: 'Umzimvubu LM',
    11: 'Winnie Mandela LM',
    12: 'Mnquma LM',
    13: 'Great Kei LM',
    14: 'Amahlathi LM',
    17: 'GPL',
    20: 'Nelson Mandela Bay MM',
    21: 'Buffalo City MM',
    25: 'GDoH',
}

# End Product Final tab grand-total row (row 30) uses =SUM() formulas.
# Restoring from backup leaves those formulas intact; the sum will now be
# a real sum of remaining numeric cells (blank-treats-as-zero for TBD text),
# which is what "work with as much as possible" asks for.


def _latest_backup(target):
    stem, ext = os.path.splitext(target)
    matches = sorted(glob.glob(f'{stem}_pre_TBD_backup_*{ext}'))
    return matches[-1] if matches else None


def restore(target):
    bak = _latest_backup(target)
    if not bak:
        raise FileNotFoundError(f'No pre-TBD backup found for {target}')
    shutil.copy2(bak, target)
    print(f'  restored: {os.path.basename(target)} <- {os.path.basename(bak)}')


def patch_sheet(path, sheet_name, batch_col_label=None):
    wb = load_workbook(path)
    ws = wb[sheet_name]

    if batch_col_label:
        actual = ws.cell(row=1, column=2).value
        assert actual == batch_col_label, f'expected B1={batch_col_label!r}, got {actual!r}'

    tbd_font = Font(italic=True, color='C00000')
    ctr = Alignment(horizontal='center')

    for row, label in PHANTOM_ROWS.items():
        actual = ws.cell(row=row, column=1).value
        if actual != label:
            print(f'  WARN: row {row} expected {label!r}, found {actual!r}; skipping')
            continue
        for col in (2, 3, 4):
            cell = ws.cell(row=row, column=col, value='TBD')
            cell.font = tbd_font
            cell.alignment = ctr
        print(f'  row {row:2d} ({label}) -> TBD')

    wb.save(path)


def main():
    print('Restoring from backups:')
    restore(EQUATION)
    restore(END_PRODUCT)

    print('\nApplying TBD to phantom rows in Display Equation (Sheet1, (M) 12/07/26 column):')
    patch_sheet(EQUATION, 'Sheet1', batch_col_label='(M) 12/07/26')

    print('\nApplying TBD to phantom rows in End Product Final tab:')
    patch_sheet(END_PRODUCT, 'Final')

    print('\nDone.')


if __name__ == '__main__':
    main()
